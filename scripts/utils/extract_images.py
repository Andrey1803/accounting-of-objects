"""
Скрипт для извлечения изображений из Excel файла akvabreg_mega.xlsx
Сохраняет изображения в static/images/catalog/ и создаёт JSON с маппингом
"""
import openpyxl
import os
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Создаём директорию
IMG_DIR = os.path.join('static', 'images', 'catalog')
os.makedirs(IMG_DIR, exist_ok=True)

wb = openpyxl.load_workbook('akvabreg_mega.xlsx', read_only=False)

# Маппинг: артикул -> путь к изображению
image_mapping = {}
# Маппинг: название листа (категории) -> список товаров с артикулами
category_mapping = {}

total_images = 0

for sheet_name in wb.sheetnames:
    if sheet_name == 'Меню':
        continue
    
    ws = wb[sheet_name]
    print(f'\nОбработка листа: {sheet_name}')
    
    # Получаем все изображения на этом листе
    images_by_row = {}  # row -> image_data
    
    if hasattr(ws, '_images') and ws._images:
        for img_idx, img in enumerate(ws._images):
            # Изображение привязано к определённой ячейке
            # Определяем к какой строке оно относится
            anchor = img.anchor
            # anchor может быть разных типов, извлекаем row
            if hasattr(anchor, 'row'):
                row_num = anchor.row
            elif hasattr(anchor, '_from'):
                row_num = anchor._from.row
            else:
                # Пробуем извлечь из строкового представления
                row_num = img_idx + 3  # Приблизительно
            
            images_by_row[row_num] = img
            total_images += 1
    
    # Читаем данные товаров (пропускаем строку 1 - пустую, строка 2 - заголовки)
    category_items = []
    for row_idx in range(3, ws.max_row + 1):
        article = ws[f'A{row_idx}'].value
        brand = ws[f'C{row_idx}'].value
        name = ws[f'E{row_idx}'].value
        
        if not article or not name:
            continue
        
        article_str = str(article).strip()
        name_str = str(name).strip()
        
        category_items.append({
            'article': article_str,
            'brand': str(brand) if brand else '',
            'name': name_str
        })
        
        # Если для этой строки есть изображение - сохраняем
        if row_num in images_by_row:
            img = images_by_row[row_num]
            # Формируем уникальное имя файла
            img_filename = f"{article_str.replace('/', '_')}.jpg"
            img_path = os.path.join(IMG_DIR, img_filename)
            
            try:
                # Сохраняем изображение
                with open(img_path, 'wb') as f:
                    f.write(img.ref.read())
                
                image_mapping[article_str] = f'images/catalog/{img_filename}'
                print(f'  ✓ Сохранено: {img_filename}')
            except Exception as e:
                print(f'  ✗ Ошибка сохранения {img_filename}: {e}')
    
    category_mapping[sheet_name] = category_items
    print(f'  Товаров на листе: {len(category_items)}')

wb.close()

# Сохраняем маппинг в JSON для последующего импорта
with open('catalog_images_mapping.json', 'w', encoding='utf-8') as f:
    json.dump({
        'images': image_mapping,
        'categories': category_mapping
    }, f, ensure_ascii=False, indent=2)

print(f'\n=== ИТОГО ===')
print(f'Изображений сохранено: {len(image_mapping)}')
print(f'Категорий обработано: {len(category_mapping)}')
print(f'Маппинг сохранён в: catalog_images_mapping.json')
