"""
Импорт категорий и материалов из Excel файла akvabreg_mega.xlsx
Создаёт иерархию категорий и связывает изображения с материалами
"""
import openpyxl
import sqlite3
import json
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

DB_PATH = 'app_data.db'
EXCEL_PATH = 'akvabreg_mega.xlsx'
MAPPING_PATH = 'catalog_images_mapping.json'

# Иконки для категорий (эмодзи)
CATEGORY_ICONS = {
    'насос': '💧',
    'труб': '🔧',
    'фитинг': '🔩',
    'кран': '🚰',
    'вентил': '🔴',
    'клапан': '⚙️',
    'фильтр': '🔬',
    'водонагреват': '🔥',
    'отоплен': '🌡️',
    'тёпл': '🏠',
    'тепл': '🏠',
    'канализац': '🚽',
    'мембран': '🫧',
    'бак': '🪣',
    'кабел': '🔌',
    'автоматик': '🤖',
    'запчаст': '🔧',
    'гайк': '🔩',
    'хомут': '🔗',
    'уплотн': '📦',
    'муфт': '🔗',
    'шланг': '🐍',
    'радиатор': '🌡️',
    'конвектор': '🌡️',
    'панел': '🖼️',
    'полотенцесушит': '🛁',
    'терм': '🌡️',
    'электр': '⚡',
    'полипропилен': '🔵',
    'обратн': '↩️',
    'манометр': '📊',
    'люк': '🚪',
    'головк': '🔲',
    'обсадн': '🔲',
    'околодец': '🕳️',
    'душ': '🚿',
    'гибк': '〰️',
    'групп': '📋',
    'демфер': '📏',
    'креплен': '📌',
    'защитн': '🛡️',
    'фольга': '✨',
    'шкаф': '🗄️',
    'коллектор': '🔀',
    'гидравл': '⚙️',
    'двигател': '⚡',
    'диффузор': '💨',
    'измельчител': '🔄',
    'комплект': '📦',
    'консол': '📐',
    'корпус': '📦',
    'крыльчатк': '🌀',
    'направл': '➡️',
    'рабоч': '⚙️',
    'сальник': '🔘',
    'ремонт': '🔧',
    'статор': '⚡',
    'ротор': '⚡',
    'штуцер': '🔗',
    'лист': '📄',
    'запасн': '📦',
    'повысит': '📈',
    'бассейн': '🏊',
    'канализац': '🚽',
    'воздушн': '💨',
    'фонтан': '⛲',
    'ручеек': '🌊',
    'скважин': '🕳️',
    'фекальн': '💩',
    'насосн': '🔧',
    'колодезн': '🕳️',
    'дренажн': '💦',
    'поверхностн': '🔧',
    'шламов': '⛏️',
    'циркуляцион': '🔄',
}

def get_icon_for_category(name):
    """Подобрать иконку для категории по названию"""
    name_lower = name.lower()
    for key, icon in CATEGORY_ICONS.items():
        if key in name_lower:
            return icon
    return '📦'  # Default icon

def get_conn():
    return sqlite3.connect(DB_PATH)

def create_image_path_table():
    """Добавить колонку image_path в catalog_materials если её нет"""
    conn = get_conn()
    try:
        conn.execute("ALTER TABLE catalog_materials ADD COLUMN image_path TEXT DEFAULT ''")
        conn.commit()
        print("✓ Добавлена колонка image_path в catalog_materials")
    except Exception as e:
        if "duplicate column" in str(e).lower() or "duplicate column name" in str(e).lower():
            print("ℹ Колонка image_path уже существует")
        else:
            print(f"⚠ Ошибка при добавлении колонки: {e}")
    finally:
        conn.close()

def import_categories_from_excel():
    """Импортировать категории из Excel файла с иерархией"""
    print("\n📂 Импорт категорий из Excel...")
    
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    conn = get_conn()
    user_id = 999  # Пользователь по умолчанию (замените на реальный ID)
    
    # Маппинг: название категории -> id в БД
    category_id_map = {}
    
    # Словарь для определения родительских категорий
    # Лист "Меню" содержит главную навигацию
    main_categories = []
    sub_categories = {}  # parent_name -> [child_names]
    
    # Сначала загружаем лист "Меню" если есть
    if 'Меню' in wb.sheetnames:
        menu_ws = wb['Меню']
        for row in menu_ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0]:
                main_categories.append(str(row[0]).strip())
    
    print(f"  Найдено главных категорий: {len(main_categories)}")
    
    # Обрабатываем все листы кроме "Меню"
    imported = 0
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Меню':
            continue
        
        ws = wb[sheet_name]
        category_name = sheet_name.strip()
        icon = get_icon_for_category(category_name)
        
        # Проверяем есть ли уже такая категория
        existing = conn.execute(
            "SELECT id FROM categories WHERE name = ? AND user_id = ?",
            (category_name, user_id)
        ).fetchone()
        
        if existing:
            category_id_map[category_name] = existing[0]
            continue
        
        # Определяем parent_id
        parent_id = None
        
        # Пытаемся определить родителя по контексту
        # Например "Бытовые насосы", "Скважинные насосы" -> родитель "Насосы"
        parent_candidates = {
            'насос': 'Насосы',
            'труб': 'Трубы',
            'фитинг': 'Фитинги',
        }
        
        for key, parent_name in parent_candidates.items():
            if key in category_name.lower():
                # Создаём или находим родителя
                parent_exists = conn.execute(
                    "SELECT id FROM categories WHERE name = ? AND user_id = ?",
                    (parent_name, user_id)
                ).fetchone()
                
                if parent_exists:
                    parent_id = parent_exists[0]
                else:
                    cur = conn.execute(
                        "INSERT INTO categories (user_id, name, parent_id, category_type) VALUES (?, ?, ?, ?)",
                        (user_id, parent_name, None, 'material')
                    )
                    conn.commit()
                    parent_id = cur.lastrowid
                    category_id_map[parent_name] = parent_id
                    print(f"  Создана главная категория: {parent_name}")
                break
        
        # Создаём категорию
        cur = conn.execute(
            "INSERT INTO categories (user_id, name, parent_id, category_type) VALUES (?, ?, ?, ?)",
            (user_id, category_name, parent_id, 'material')
        )
        conn.commit()
        cat_id = cur.lastrowid
        category_id_map[category_name] = cat_id
        imported += 1
    
    wb.close()
    conn.close()
    
    print(f"✓ Импортировано категорий: {imported}")
    return category_id_map

def link_images_to_materials():
    """Связать изображения с материалами по артикулу"""
    print("\n🖼️  Связывание изображений с материалами...")
    
    if not os.path.exists(MAPPING_PATH):
        print("⚠ Файл маппинга не найден, пропускаем")
        return
    
    with open(MAPPING_PATH, 'r', encoding='utf-8') as f:
        mapping = json.load(f)
    
    image_map = mapping.get('images', {})
    conn = get_conn()
    user_id = 999
    
    linked_count = 0
    
    # Для каждого материала ищем изображение по артикулу
    materials = conn.execute(
        "SELECT id, name FROM catalog_materials WHERE user_id = ?",
        (user_id,)
    ).fetchall()
    
    for mat_id, mat_name in materials:
        # Пытаемся найти материал в Excel по названию
        for article, img_path in image_map.items():
            if mat_name and article in mat_name:
                conn.execute(
                    "UPDATE catalog_materials SET image_path = ? WHERE id = ?",
                    (img_path, mat_id)
                )
                linked_count += 1
                break
    
    conn.commit()
    conn.close()
    
    print(f"✓ Связано изображений: {linked_count}")

def create_default_user():
    """Создать пользователя с id=999 если его нет"""
    conn = get_conn()
    user = conn.execute("SELECT id FROM users WHERE id = 999").fetchone()
    if not user:
        from auth import hash_pw
        conn.execute(
            "INSERT INTO users (id, username, password_hash, role, email, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (999, 'demo', hash_pw('demo'), 'user', 'demo@example.com', '')
        )
        conn.commit()
        print("✓ Создан демо-пользователь (id=999, login: demo, password: demo)")
    conn.close()

if __name__ == '__main__':
    print("=" * 60)
    print("ИМПОРТ КАТЕГОРИЙ И МАТЕРИАЛОВ ИЗ EXCEL")
    print("=" * 60)
    
    # 1. Создаём пользователя
    create_default_user()
    
    # 2. Добавляем колонку image_path
    create_image_path_table()
    
    # 3. Импортируем категории
    category_id_map = import_categories_from_excel()
    
    # 4. Связываем изображения
    link_images_to_materials()
    
    print("\n" + "=" * 60)
    print("ИМПОРТ ЗАВЕРШЁН!")
    print("=" * 60)
