# -*- coding: utf-8 -*-
"""Импорт каталога из Excel — конвертация cp1251 -> utf-8"""
import openpyxl
import os
import sys
import sqlite3

EST_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'app_data.db')
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'akvabreg_mega.xlsx')

def decode_cp1251(v):
    """openpyxl читает xlsx правильно — просто возвращаем строку"""
    if v is None: return ''
    return str(v).strip()

def main():
    if not os.path.exists(EXCEL_PATH):
        print("[ERROR] Файл не найден:", EXCEL_PATH)
        return
    if not os.path.exists(EST_DB_PATH):
        print("[ERROR] БД не найдена:", EST_DB_PATH)
        return

    print("Чтение Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    conn = sqlite3.connect(EST_DB_PATH)
    c = conn.cursor()

    # Добавляем колонки если их нет
    for col_name in ['brand', 'item_type', 'article']:
        try:
            c.execute(f"ALTER TABLE catalog_materials ADD COLUMN {col_name} TEXT DEFAULT ''")
            print(f"  Добавлена колонка: {col_name}")
        except:
            pass  # Колонка уже есть

    USER_ID = int(sys.argv[1]) if len(sys.argv) > 1 else 999

    c.execute("DELETE FROM catalog_materials WHERE user_id = ?", (USER_ID,))
    conn.commit()

    skip_sheets = {"Меню"}
    total = 0
    skipped_total = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Ищем заголовки
        hdr_idx = None
        for i, row in enumerate(rows[:5]):
            vals = [_raw(v).lower() for v in row if v is not None]
            raw_line = ' '.join(vals)
            if any(kw in raw_line for kw in ['ртикул', 'именование', 'аименование', 'наимен', 'назв', 'nazwa', 'name']):
                hdr_idx = i
                break
        if hdr_idx is None:
            continue

        # Маппинг колонок (akvabreg_mega: РРЦ/розница, опт1, опт2 — в оптовую цену только опт2)
        hdr = [_s(v).lower() for v in rows[hdr_idx]]
        col = {}
        for ci, h in enumerate(hdr):
            if 'именование' in h or 'аименование' in h or 'наимен' in h or 'назв' in h or 'nazwa' in h:
                col['name'] = ci
            elif 'опис' in h or 'писан' in h or 'писание' in h:
                col['desc'] = ci
            elif 'опт' in h and '2' in h:
                col['opt2'] = ci
            elif 'опт' in h and '1' in h:
                col['opt1'] = ci
            elif 'ррц' in h or 'розница' in h or 'рознич' in h or 'retail' in h:
                col['retail'] = ci
            elif 'закуп' in h or 'purchase' in h:
                col['purchase'] = ci
            elif ('цена' in h or 'стоим' in h) and 'опт' not in h and 'закуп' not in h and 'дилер' not in h:
                if 'retail' not in col:
                    col['retail'] = ci
            elif 'ед' in h or 'unit' in h:
                col['unit'] = ci
            elif 'категори' in h or 'атегори' in h or 'category' in h:
                col['category'] = ci
            elif 'бренд' in h:
                col['brand'] = ci
            elif 'тип' in h:
                col['item_type'] = ci
            elif 'ртикул' in h or 'артикул' in h:
                col['article'] = ci

        for k, dv in [('name', 4), ('desc', 5), ('retail', -1)]:
            if k not in col:
                col[k] = dv
        if 'unit' not in col:
            col['unit'] = -1
        if 'category' not in col:
            col['category'] = -1
        if 'brand' not in col:
            col['brand'] = -1
        if 'item_type' not in col:
            col['item_type'] = -1
        if 'article' not in col:
            col['article'] = -1
        if 'opt1' not in col:
            col['opt1'] = -1
        if 'opt2' not in col:
            col['opt2'] = -1
        if 'purchase' not in col:
            col['purchase'] = -1

        added = 0
        skipped = 0
        for row in rows[hdr_idx+1:]:
            if not row or all(v is None for v in row):
                continue

            name = _c(row, col['name'])
            if not name or len(name) < 3:
                continue

            non_empty = sum(1 for v in row if v is not None and _s(v).strip())
            if non_empty <= 1:
                continue

            desc   = _c(row, col.get('desc', -1))
            unit   = _c(row, col.get('unit', -1)) or 'шт'
            retail = _p(row, col.get('retail', -1))
            if retail <= 0:
                skipped += 1
                continue

            opt2 = _p(row, col.get('opt2', -1))
            # Оптовая цена в БД — только из колонки «опт 2»; если пусто — как раньше, розница
            wholesale = opt2 if opt2 > 0 else retail

            purchase = _p(row, col.get('purchase', -1))
            if purchase <= 0:
                purchase = round(retail * 0.7, 2)
            # Категория — имя листа
            category = decode_cp1251(sheet_name)
            # Бренд и тип из колонок
            brand = _c(row, col.get('brand', -1))
            item_type = _c(row, col.get('item_type', -1))
            article = _c(row, col.get('article', -1))

            try:
                c.execute("""
                    INSERT INTO catalog_materials
                    (user_id, name, unit, category, purchase_price, retail_price, wholesale_price,
                     description, min_wholesale_qty, use_count, brand, item_type, article)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?)
                """, (USER_ID, name, unit, category, purchase, retail, wholesale, desc, 10, brand or '', item_type or '', article or ''))
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1

        if added:
            print("  [{}]: +{} (пропущено: {})".format(decode_cp1251(sheet_name), added, skipped))
        total += added
        skipped_total += skipped

    conn.commit()
    conn.close()
    print("\nДобавлено: {}, Пропущено: {} (user_id={})".format(total, skipped_total, USER_ID))

def _raw(v):
    """Сырая строка для поиска — без декодирования"""
    if v is None: return ''
    return str(v).strip().lower()

def _s(v):
    if v is None: return ''
    return decode_cp1251(v)

def _c(row, idx):
    if idx < len(row) and row[idx] is not None:
        return decode_cp1251(row[idx])
    return ''

def _p(row, idx):
    if idx >= len(row) or row[idx] is None: return 0
    try:
        return float(str(row[idx]).replace(',','.').replace(' ','').replace('\xa0',''))
    except: return 0

if __name__ == '__main__':
    print("="*60)
    print("  Импорт — cp1251 -> utf-8")
    print("="*60)
    main()
