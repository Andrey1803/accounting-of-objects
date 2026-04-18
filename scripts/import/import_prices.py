# -*- coding: utf-8 -*-
"""Импорт цен из Excel в catalog_materials"""
import openpyxl
import sqlite3
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = 'akvabreg_mega.xlsx'
DB_PATH = 'app_data.db'

SKIP_SHEETS = {'Меню'}

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    total_imported = 0
    total_skipped = 0
    total_no_match = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Заголовки на строке 2 (индекс 1)
        hdr_idx = None
        for i, row in enumerate(rows[:5]):
            vals = [str(v).strip().lower() for v in row if v is not None]
            raw = ' '.join(vals)
            if any(kw in raw for kw in ['ртикул', 'аименование', 'именование', 'наимен', 'назв']):
                hdr_idx = i
                break

        if hdr_idx is None:
            print(f"  [{sheet_name}] Нет заголовков, пропускаю")
            continue

        # Маппинг колонок
        hdr = [str(v).strip().lower() for v in rows[hdr_idx]]
        col = {}
        for ci, h in enumerate(hdr):
            if any(kw in h for kw in ['аименование', 'именование', 'наимен', 'назв']): col['name'] = ci
            elif 'опис' in h or 'писание' in h: col['desc'] = ci
            elif 'ррц' in h or 'розница' in h or 'retail' in h: col['retail'] = ci
            elif 'опт' in h and '1' in h: col['opt1'] = ci
            elif 'опт' in h and '2' in h: col['opt2'] = ci
            elif 'ед' in h or 'unit' in h: col['unit'] = ci
            elif 'категори' in h: col['category'] = ci
            elif 'артикул' in h: col['article'] = ci
            elif h in ('тип', 'type'): col['item_type'] = ci
            elif h in ('бренд', 'brand'): col['brand'] = ci

        # Defaults
        for k, dv in [('name',0),('retail',2)]:
            if k not in col: col[k] = dv
        if 'opt1' not in col: col['opt1'] = col.get('retail', 2)
        if 'opt2' not in col: col['opt2'] = col.get('retail', 2)
        if 'unit' not in col: col['unit'] = 99
        if 'category' not in col: col['category'] = -1

        added = 0
        updated = 0
        skipped = 0

        # Пропускаем строки заголовков и маркеры категорий
        data_start = hdr_idx + 1
        for row in rows[data_start:]:
            if not row or all(v is None for v in row):
                continue

            non_empty = sum(1 for v in row if v is not None and str(v).strip())
            if non_empty <= 1:
                continue

            name = _c(row, col['name'])
            if not name or len(name) < 3:
                continue

            retail = _p(row, col['retail'])
            if retail <= 0:
                skipped += 1
                continue

            purchase = round(retail * 0.7, 2)
            opt1 = _p(row, col['opt1'])
            opt2 = _p(row, col['opt2'])
            unit = _c(row, col.get('unit', 99)) or 'шт'
            desc = _c(row, col.get('desc', 1))
            article = _c(row, col.get('article', 0))
            brand = _c(row, col.get('brand', 0))
            item_type = _c(row, col.get('item_type', 0))
            category = _c(row, col.get('category', -1)) or sheet_name

            # Ищем по артикулу или имени
            if article:
                c.execute("SELECT id FROM catalog_materials WHERE user_id=1 AND (article=? OR name=?)", (article, name))
            else:
                c.execute("SELECT id FROM catalog_materials WHERE user_id=1 AND name=?", (name,))

            existing = c.fetchone()
            if existing:
                c.execute("""UPDATE catalog_materials SET
                    purchase_price=?, retail_price=?, wholesale_price=?,
                    unit=?, description=?, article=?, brand=?, item_type=?, opt1=?, opt2=?
                    WHERE id=?""",
                    (purchase, retail, opt2, unit, desc, article, brand, item_type, opt1, opt2, existing[0]))
                updated += 1
            else:
                c.execute("""INSERT INTO catalog_materials
                    (user_id, name, unit, category, purchase_price, retail_price, wholesale_price,
                     description, min_wholesale_qty, use_count, article, brand, item_type, opt1, opt2)
                    VALUES (1, ?, ?, ?, ?, ?, ?, ?, 10, 0, ?, ?, ?, ?, ?)""",
                    (name, unit, category, purchase, retail, opt2, desc, article, brand, item_type, opt1, opt2))
                added += 1

        if added or updated:
            print(f"  [{sheet_name}]: +{added} новых, ~{updated} обновлено (пропущено: {skipped})")
        total_imported += added + updated
        total_skipped += skipped

    conn.commit()
    conn.close()
    wb.close()
    print(f"\nИТОГО: Импортировано={total_imported}, Пропущено (без цен)={total_skipped}")

def _c(row, idx):
    if idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return ''

def _p(row, idx):
    if idx >= len(row) or row[idx] is None: return 0
    try:
        return float(str(row[idx]).replace(',','.').replace(' ','').replace('\xa0',''))
    except: return 0

if __name__ == '__main__':
    print("=" * 60)
    print("  Импорт цен из Excel")
    print("=" * 60)
    main()
