# -*- coding: utf-8 -*-
"""Поиск IRCEM в Excel и импорт если найден"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl, sqlite3

wb = openpyxl.load_workbook('akvabreg_mega.xlsx', data_only=True)

found_count = 0
for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    count = 0
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and 'IRCEM' in str(cell).upper():
                count += 1
                break
    if count > 0:
        print(f'{ws_name}: {count} товаров с IRCEM')
        found_count += count

print(f'\nИтого в Excel: {found_count} товаров IRCEM')
wb.close()

if found_count > 0:
    print('\nIRCEM найден в Excel, но НЕ импортирован в БД.')
    print('Нужно запустить импорт цен ещё раз для недостающих товаров.')
