import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('akvabreg_mega.xlsx', read_only=False)
print(f'Листов: {len(wb.sheetnames)}')

# Смотрим лист "Бытовые насосы"
ws = wb['Бытовые насосы']
print(f'\nЛист: {ws.title}')

print('\n=== ПЕРВЫЕ 5 СТРОК (колонки A-F) ===')
for i in range(1, 6):
    row_data = []
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        val = ws[f'{col_letter}{i}'].value
        row_data.append(f'{col_letter}={val}')
    print(f'Строка {i}: {row_data}')

# Проверяем есть ли изображения
print('\n=== ПРОВЕРКА ИЗОБРАЖЕНИЙ ===')
if hasattr(ws, '_images'):
    print(f'Изображений: {len(ws._images)}')
    if ws._images:
        img = ws._images[0]
        print(f'Первое изображение: {img.ref}')
else:
    print('Нет изображений')

wb.close()
