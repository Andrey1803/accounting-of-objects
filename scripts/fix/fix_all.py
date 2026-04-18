"""
Комплексное исправление проекта:
1. Очистка дубликатов каталога
2. Импорт для правильного пользователя (user_id=1)
3. Связывание фото по артикулу
"""
import sqlite3
import json
import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

DB_PATH = 'app_data.db'
EXCEL_PATH = 'akvabreg_mega.xlsx'
MAPPING_PATH = 'catalog_images_mapping.json'
USER_ID = 1  # Основной пользователь

def get_conn():
    return sqlite3.connect(DB_PATH)

# ============================================================
# ШАГ 1: Очистка старых данных
# ============================================================
def cleanup():
    print("\n" + "=" * 60)
    print("ШАГ 1: ОЧИСТКА ДУБЛИКАТОВ")
    print("=" * 60)
    
    conn = get_conn()
    
    # Считаем что удаляем
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM catalog_materials WHERE user_id = ?", (USER_ID,))
    old_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM categories WHERE user_id = ?", (USER_ID,))
    old_cats = c.fetchone()[0]
    
    print(f"  Материалов для user_id={USER_ID}: {old_count}")
    print(f"  Категорий для user_id={USER_ID}: {old_cats}")
    
    # Удаляем старые данные
    conn.execute("DELETE FROM catalog_materials WHERE user_id = ?", (USER_ID,))
    conn.execute("DELETE FROM categories WHERE user_id = ?", (USER_ID,))
    conn.commit()
    
    print(f"  ✅ Удалено {old_count} материалов и {old_cats} категорий")
    conn.close()

# ============================================================
# ШАГ 2: Импорт категорий из Excel
# ============================================================
def import_categories():
    print("\n" + "=" * 60)
    print("ШАГ 2: ИМПОРТ КАТЕГОРИЙ")
    print("=" * 60)
    
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    conn = get_conn()
    
    # Иконки для категорий
    CATEGORY_ICONS = {
        'насос': '💧', 'труб': '🔧', 'фитинг': '🔩', 'кран': '🚰', 'вентил': '🔴',
        'клапан': '⚙️', 'фильтр': '🔬', 'водонагреват': '🔥', 'отоплен': '🌡️',
        'тёпл': '🏠', 'тепл': '🏠', 'канализац': '🚽', 'мембран': '🫧', 'бак': '🪣',
        'кабел': '🔌', 'автоматик': '🤖', 'запчаст': '🔧', 'муфт': '🔗', 'шланг': '🐍',
        'радиатор': '🌡️', 'полипропилен': '🔵', 'электр': '⚡', 'манометр': '📊',
        'люк': '🚪', 'душ': '🚿', 'гибк': '〰️', 'шкаф': '🗄️', 'коллектор': '🔀',
        'гидравл': '⚙️', 'двигател': '⚡', 'креплен': '📌', 'демпфер': '📏',
        'фольга': '✨', 'защитн': '🛡️', 'уплотн': '📦', 'комплект': '📦',
        'запасн': '📦', 'повысит': '📈', 'бассейн': '🏊', 'воздушн': '💨',
        'фонтан': '⛲', 'ручеек': '🌊', 'скважин': '🕳️', 'фекальн': '💩',
        'дренажн': '💦', 'поверхностн': '🔧', 'циркуляцион': '🔄', 'конвектор': '🌡️',
    }
    
    def get_icon(name):
        n = name.lower()
        for key, icon in CATEGORY_ICONS.items():
            if key in n:
                return icon
        return '📦'
    
    # Определяем родительские категории
    parent_groups = {
        'Насосы': ['бытовые', 'скважинные', 'фекальные', 'насосные станции', 'колодезные', 
                   'дренажные', 'поверхностные', 'шламовые', 'циркуляционные', 'повысительные',
                   'бассейн', 'канализационные', 'воздушные', 'фонтан', 'ручеек'],
        'Трубы': ['труба', 'обсадн', 'питьевая'],
        'Фитинги': ['фитинг', 'муфт', 'штуцер'],
        'Запчасти': ['запасн', 'гидравлическ', 'двигател', 'диффузор', 'измельчител',
                    'комплект гайк', 'консоли', 'корпус', 'крыльчатк', 'направл',
                    'рабоч', 'сальник', 'ремонт', 'статор', 'ротор'],
        'Фильтры': ['фильтр', 'картридж', 'колб', 'питьев', 'осмос', 'грязевик'],
        'Отопление': ['отоплен', 'радиатор', 'конвектор', 'тёпл', 'тепл', 'демпфер',
                     'коллекторн', 'шкаф', 'фольга'],
        'Водонагреватели': ['водонагреват', 'электрическ', 'газов'],
        'Полипропилен': ['полипропилен', 'комбинированн'],
        'Краны': ['кран'],
        'Шланги': ['шланг', 'всасывающ'],
        'Кабели': ['кабел'],
        'Автоматика': ['автоматик', 'трубн'],
        'Мембранные баки': ['мембранн', 'гидроаккумулятор', 'кронштейн', 'мембраны для баков', 'фланцы'],
        'Расширительные баки': ['расширительн'],
        'Канализация': ['канализац', 'люк'],
        'Обратные клапана': ['обратн'],
        'Манометры': ['манометр'],
        'Хомуты': ['хомут'],
        'Уплотнители': ['уплотн'],
        'Душ': ['душ'],
        'Комплектующие': ['комплектующ'],
        'Водоснабжение': ['адаптер', 'оголовк', 'трос', 'защитн'],
    }
    
    category_id_map = {}
    imported_cats = 0
    
    # Создаём родительские категории
    for parent_name in parent_groups.keys():
        icon = get_icon(parent_name)
        try:
            cur = conn.execute(
                "INSERT OR IGNORE INTO categories (user_id, name, parent_id, category_type) VALUES (?, ?, NULL, ?)",
                (USER_ID, parent_name, 'material')
            )
            conn.commit()
            # Получаем ID
            existing = conn.execute(
                "SELECT id FROM categories WHERE user_id = ? AND name = ?",
                (USER_ID, parent_name)
            ).fetchone()
            if existing:
                category_id_map[parent_name] = existing[0]
                imported_cats += 1
        except Exception as e:
            print(f"  ⚠ Ошибка создания категории {parent_name}: {e}")
    
    # Создаём подкатегории из листов Excel
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Меню':
            continue
        
        # Определяем родителя
        parent_id = None
        sheet_lower = sheet_name.lower()
        
        for parent_name, keywords in parent_groups.items():
            for kw in keywords:
                if kw in sheet_lower:
                    parent_id = category_id_map.get(parent_name)
                    break
            if parent_id:
                break
        
        cat_name = sheet_name.strip()
        try:
            cur = conn.execute(
                "INSERT OR IGNORE INTO categories (user_id, name, parent_id, category_type) VALUES (?, ?, ?, ?)",
                (USER_ID, cat_name, parent_id, 'material')
            )
            conn.commit()
            existing = conn.execute(
                "SELECT id FROM categories WHERE user_id = ? AND name = ?",
                (USER_ID, cat_name)
            ).fetchone()
            if existing:
                category_id_map[cat_name] = existing[0]
                imported_cats += 1
        except Exception as e:
            print(f"  ⚠ Ошибка создания подкатегории {cat_name}: {e}")
    
    wb.close()
    conn.close()
    
    print(f"  ✅ Импортировано категорий: {imported_cats}")
    print(f"  Родительских: {len(parent_groups)}")
    print(f"  Подкатегорий: {imported_cats - len(parent_groups)}")

# ============================================================
# ШАГ 3: Импорт материалов из Excel
# ============================================================
def import_materials():
    print("\n" + "=" * 60)
    print("ШАГ 3: ИМПОРТ МАТЕРИАЛОВ")
    print("=" * 60)
    
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    conn = get_conn()
    
    total_imported = 0
    total_skipped = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Меню':
            continue
        
        ws = wb[sheet_name]
        category = sheet_name.strip()
        row_count = 0
        
        # Пропускаем первые 2 строки (пустая + заголовки)
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or not row[4]:  # row[4] = Наименование
                continue
            
            article = str(row[0]).strip() if row[0] else ''
            brand = str(row[2]).strip() if row[2] else ''
            item_type = str(row[3]).strip() if row[3] else ''
            name = str(row[4]).strip() if row[4] else ''
            description = str(row[5]).strip() if row[5] else ''
            
            if not name:
                continue
            
            # Парсим цены из описания (если есть)
            purchase_price = 0
            retail_price = 0
            
            # Проверяем дубликаты
            exists = conn.execute(
                "SELECT id FROM catalog_materials WHERE user_id = ? AND article = ? AND article != ''",
                (USER_ID, article)
            ).fetchone()
            
            if exists:
                total_skipped += 1
                continue
            
            # Если артикула нет - проверяем по названию
            if not article:
                exists = conn.execute(
                    "SELECT id FROM catalog_materials WHERE user_id = ? AND name = ?",
                    (USER_ID, name)
                ).fetchone()
                if exists:
                    total_skipped += 1
                    continue
            
            # Вставляем материал
            try:
                conn.execute("""
                    INSERT INTO catalog_materials 
                    (user_id, article, brand, name, unit, category, item_type, 
                     purchase_price, retail_price, wholesale_price, min_wholesale_qty, 
                     description, use_count, image_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    USER_ID,
                    article,
                    brand,
                    name,
                    'шт',  # unit по умолчанию
                    category,
                    item_type,
                    purchase_price,
                    retail_price,
                    retail_price,  # wholesale = retail
                    10,  # min_wholesale_qty
                    description,
                    1,  # use_count
                    ''  # image_path (заполним позже)
                ))
                row_count += 1
                total_imported += 1
            except Exception as e:
                print(f"  ⚠ Ошибка импорта: {name[:50]} - {e}")
                total_skipped += 1
        
        if row_count > 0:
            print(f"  {category}: {row_count}")
        
        if total_imported > 0 and total_imported % 500 == 0:
            conn.commit()
            print(f"  ... уже импортировано: {total_imported}")
    
    conn.commit()
    wb.close()
    conn.close()
    
    print(f"\n  ✅ Импортировано материалов: {total_imported}")
    print(f"  ⏭ Пропущено дубликатов: {total_skipped}")

# ============================================================
# ШАГ 4: Связывание фото с материалами
# ============================================================
def link_images():
    print("\n" + "=" * 60)
    print("ШАГ 4: СВЯЗЫВАНИЕ ФОТО")
    print("=" * 60)
    
    if not os.path.exists(MAPPING_PATH):
        print("  ⚠ Файл маппинга не найден")
        return
    
    with open(MAPPING_PATH, 'r', encoding='utf-8') as f:
        mapping = json.load(f)
    
    image_map = mapping.get('images', {})
    print(f"  Загружено изображений из маппинга: {len(image_map)}")
    
    conn = get_conn()
    linked = 0
    
    materials = conn.execute(
        "SELECT id, article FROM catalog_materials WHERE user_id = ?",
        (USER_ID,)
    ).fetchall()
    
    for mat_id, article in materials:
        if article and article in image_map:
            img_path = image_map[article]
            # Проверяем что файл существует
            full_path = os.path.join('static', img_path)
            if os.path.exists(full_path) and os.path.getsize(full_path) > 0:
                conn.execute(
                    "UPDATE catalog_materials SET image_path = ? WHERE id = ?",
                    (img_path, mat_id)
                )
                linked += 1
    
    conn.commit()
    conn.close()
    
    print(f"  ✅ Связано фото: {linked}")

# ============================================================
# ШАГ 5: Финальная статистика
# ============================================================
def final_stats():
    print("\n" + "=" * 60)
    print("ШАГ 5: ФИНАЛЬНАЯ СТАТИСТИКА")
    print("=" * 60)
    
    conn = get_conn()
    c = conn.cursor()
    
    c.execute("SELECT COUNT(*) FROM catalog_materials WHERE user_id = ?", (USER_ID,))
    print(f"  Материалов: {c.fetchone()[0]}")
    
    c.execute("SELECT COUNT(*) FROM categories WHERE user_id = ?", (USER_ID,))
    print(f"  Категорий: {c.fetchone()[0]}")
    
    c.execute("SELECT COUNT(*) FROM catalog_materials WHERE user_id = ? AND image_path IS NOT NULL AND image_path != ''", (USER_ID,))
    with_photo = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM catalog_materials WHERE user_id = ?", (USER_ID,))
    total = c.fetchone()[0]
    print(f"  С фото: {with_photo}/{total} ({with_photo*100//total if total > 0 else 0}%)")
    
    c.execute("SELECT COUNT(DISTINCT category) FROM catalog_materials WHERE user_id = ?", (USER_ID,))
    print(f"  Уникальных категорий в материалах: {c.fetchone()[0]}")
    
    conn.close()

# ============================================================
# ЗАПУСК
# ============================================================
if __name__ == '__main__':
    print("\n" + "=" * 60)
    print("КОМПЛЕКСНОЕ ИСПРАВЛЕНИЕ ПРОЕКТА")
    print(f"Целевой пользователь: user_id={USER_ID}")
    print("=" * 60)
    
    cleanup()
    import_categories()
    import_materials()
    link_images()
    final_stats()
    
    print("\n" + "=" * 60)
    print("✅ ВСЕ ОПЕРАЦИИ ЗАВЕРШЕНЫ!")
    print("=" * 60)
