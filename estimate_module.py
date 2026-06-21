"""
Модуль для работы со сметами, материалами и работами.
"""
from flask import Blueprint, request, jsonify, render_template, send_file, session
from flask_login import current_user, login_required
from datetime import datetime
from database import fetch_all, fetch_one, execute, execute_rowcount
import logging
import re
import openpyxl
import os
import json
import urllib.request
import urllib.parse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

logger = logging.getLogger(__name__)

# --- Импорт позиций из PDF (parse-pdf): фильтры строк и сопоставление с каталогом ---
_PDF_HEADER_START = re.compile(
    r'^(?:\s*№\s|товар|наимен|наименование|работ|услуг|коли|кол-во|количество|ед\.?\s*изм|'
    r'цена|стоимость|срок|артикул|н/п|№п/п|поз\.?)',
    re.I,
)
_PDF_TOTALS_ROW = re.compile(
    r'\b(?:итого|всего|в\s*том\s*числе|в\s*т\.?\s*ч\.?|ндс|сумма\s*ндс|скидк|к\s*оплате|'
    r'прописью|страниц|листов|документ\s+составлен|без\s+ндс)\b',
    re.I,
)
_PDF_NON_PRODUCT_ROW = re.compile(
    r'(?:'
    r'\b(?:январ|феврал|март|апрел|мая|июн|июл|август|сентябр|октябр|ноябр|декабр)\w*\s+\d{4}'
    r'|\b\d{1,2}\s+(?:январ|феврал|март|апрел|мая|июн|июл|август|сентябр|октябр|ноябр|декабр)'
    r'|\b(?:г\.?\s*ооо|унп\s*\d|р\/с\s*\d)'
    r')',
    re.I,
)
_PDF_STOPWORDS = frozenset({
    'для', 'без', 'или', 'это', 'все', 'всех', 'тип', 'вид', 'как', 'при', 'под', 'над',
    'размер', 'цвет', 'белым', 'белый', 'черным', 'черный', 'серый', 'коричн',
    'штук', 'упак', 'короб', 'масса', 'вес', 'год', 'мес',
})
# Корзина ИМ (akvabreg.by и аналоги): «цена руб. N шт сумма руб.» + «Артикул …»
_PDF_CART_PRICE_LINE = re.compile(
    r'^(\d+(?:[.,]\d+)?)\s*руб\.\s+(\d+(?:[.,]\d+)?)\s*шт\s+(\d+(?:[.,]\d+)?)\s*руб\.?\s*$',
    re.I,
)
_PDF_CART_ARTICLE_LINE = re.compile(r'^Артикул\s+(\d+)\s*$', re.I)
# Оптовый счёт в PDF: закуп часто указан без НДС; в каталог пишем purchase с НДС (20 %).
_PDF_WHOLESALE_EX_VAT_TO_WITH_VAT = 1.2


def _pdf_parse_cart_ready_format(full_text):
    """
    PDF «Готовые к заказу»: наименование (в т.ч. несколько строк), строка с ценой/кол-вом/суммой,
    затем «Артикул». Возвращает строки [name, unit_price, qty, total, article] для api_parse_pdf.
    """
    if not full_text or 'руб.' not in full_text.lower():
        return []

    raw_lines = []
    for raw in full_text.replace('\r\n', '\n').replace('\r', '\n').split('\n'):
        s = ' '.join(raw.split()).strip()
        if not s:
            continue
        low = s.lower()
        if 'image not found' in low or low == 'type unknown':
            continue
        if re.match(r'^--\s*\d+\s+of\s+\d+\s*--$', s, re.I):
            continue
        if low.startswith('готовые к заказу'):
            continue
        raw_lines.append(s)

    def _footer(line):
        lo = line.lower()
        if re.match(r'^итого\s*:', lo):
            return True
        if lo.startswith('интернет-магазин') and len(line) < 90:
            return True
        if line.startswith('http://') or line.startswith('https://'):
            return True
        if re.search(r'^\+\d', line) and re.search(r'\(\d', line):
            return True
        if 'info@' in lo:
            return True
        if re.search(r'ул\.|минская область|щомыслицк', lo) and len(line) > 35:
            return True
        return False

    lines = raw_lines
    n = len(lines)
    out = []
    i = 0
    while i < n:
        line = lines[i]
        if _footer(line):
            break
        if _PDF_CART_PRICE_LINE.match(line):
            i += 1
            continue
        if _PDF_CART_ARTICLE_LINE.match(line):
            i += 1
            continue
        lo = line.lower()
        if lo.startswith('тип цены') or lo.startswith('наличие'):
            i += 1
            continue

        name_parts = []
        start_i = i
        while i < n:
            s = lines[i]
            if _footer(s):
                return out
            if _PDF_CART_PRICE_LINE.match(s):
                break
            if _PDF_CART_ARTICLE_LINE.match(s):
                i += 1
                continue
            lo2 = s.lower()
            if lo2.startswith('тип цены') or lo2.startswith('наличие'):
                i += 1
                continue
            name_parts.append(s)
            i += 1

        if i >= n:
            break
        m = _PDF_CART_PRICE_LINE.match(lines[i])
        if not m or not name_parts:
            i = start_i + 1
            continue

        unit_p, qty, tot = m.group(1), m.group(2), m.group(3)
        i += 1
        while i < n:
            lo3 = lines[i].lower()
            if lo3.startswith('тип цены') or lo3.startswith('наличие'):
                i += 1
                continue
            break

        art = ''
        if i < n and _PDF_CART_ARTICLE_LINE.match(lines[i]):
            art = _PDF_CART_ARTICLE_LINE.match(lines[i]).group(1)
            i += 1

        name = ' '.join(name_parts).strip()
        if name:
            out.append(
                [
                    name,
                    unit_p.replace(',', '.'),
                    qty.replace(',', '.'),
                    tot.replace(',', '.'),
                    art,
                ]
            )
    return out


def _pdf_is_non_product_row(name, row_text=''):
    """Строки подвала счёта (дата, реквизиты ООО), не номенклатура."""
    n = (name or '').strip()
    blob = f'{n} {(row_text or "").strip()}'.lower()
    if not blob.strip():
        return True
    if _PDF_TOTALS_ROW.search(blob):
        return True
    if len(n) < 40 and _PDF_NON_PRODUCT_ROW.search(n):
        return True
    if re.search(r'\d{4}\s*г\.?', n, re.I) and re.search(r'\b(?:ооо|зао|чуп)\b', n, re.I):
        return True
    if len(n) < 50 and re.search(
        r'\b(?:январ|феврал|март|апрел|мая|июн|июл|август|сентябр|октябр|ноябр|декабр)\w*\b',
        n,
        re.I,
    ) and re.search(r'\d{4}', n):
        return True
    if len(n) < 18 and re.fullmatch(r'(?:ооо|зао|оао|чуп|ип)\b.*', n, re.I):
        return True
    return False


def _pdf_substantial_table_rows(all_rows, import_mode='retail'):
    """Есть ли в PDF нормальная таблица сметы (заголовок + несколько строк данных)."""
    all_rows = _pdf_filter_table_rows(all_rows)
    if not all_rows:
        return False
    header = _pdf_detect_header_layout(all_rows, import_mode)
    if not header:
        return False
    data_like = 0
    for row in all_rows:
        if not row or len(row) < 4:
            continue
        cells = [str(c or '').strip() for c in row]
        ni = header.get('name_idx')
        if ni is None or int(ni) >= len(cells):
            continue
        name = cells[int(ni)]
        if len(name) < 3 or _pdf_is_non_product_row(name, ' '.join(cells)):
            continue
        if _PDF_HEADER_START.search(name.lower()):
            continue
        data_like += 1
        if data_like >= 3:
            return True
    return False


def _pdf_row_product_title(raw):
    """Выделить наименование из типичной строки прайса (№, цена в начале, шт/м/пог.м, суммы, страна)."""
    s = ' '.join(str(raw).split()).strip()
    if not s:
        return ''
    s = re.sub(r'^\d+\s+', '', s)
    s = re.sub(r'^\d{1,3}(?:\s+\d{3})+[\.,]\d{2}\s+', '', s)
    s = re.sub(r'^\d+[\.,]\d{2}\s+', '', s)
    s = re.sub(r'\s*В\s*наличии.*$', '', s, flags=re.I)
    s = re.sub(r'\s*Под\s*заказ.*$', '', s, flags=re.I)
    s = re.sub(r'\s*\([^)]*уверенность[^)]*\)\s*$', '', s, flags=re.I)
    s = re.sub(r'\s*\([^)]*неоднознач[^)]*\)\s*$', '', s, flags=re.I)

    def _cut(pattern, min_left=12):
        nonlocal s
        m = re.search(pattern, s, re.I)
        if m and m.start() >= min_left:
            s = s[: m.start()].strip()

    _price_tail = r'(?:\d{1,3}(?:\s\d{3})+[\.,]\d{2}|\d+[\.,]\d{2})'
    _cut(rf'\s+\d+(?:[.,]\d+)?\s+(?:шт\.?|шт)\s+{_price_tail}')
    _cut(rf'\s+\d+(?:[.,]\d+)?\s+пог\.?\s*м\s+{_price_tail}')
    _cut(rf'\s+\d+(?:[.,]\d+)?\s+компл\.?\s+{_price_tail}')
    _cut(rf'\s+\d+(?:[.,]\d+)?\s+м\s+{_price_tail}', min_left=18)
    _cut(rf'\s+\d+(?:[.,]\d+)?\s+м\s+{_price_tail}\s+{_price_tail}', min_left=18)

    s = re.sub(rf'\s+{_price_tail}\s+{_price_tail}\s*$', '', s)
    s = re.sub(
        r'[,]?\s*(Республика\s+Польша|Италия|Китай|Беларусь|БЕЛАРУСЬ|Пр-во\s*РБ|Пр-во\s*Рб)\s*$',
        '',
        s,
        flags=re.I,
    )
    return s.strip()


def _pdf_cell_is_unit(cell):
    """Ячейка «ед. изм.» как в печатной смете (шт / м / компл …)."""
    if cell is None:
        return False
    u = str(cell).strip().lower().replace('\xa0', '')
    u_compact = re.sub(r'\s+', '', u).rstrip('.')
    if u_compact in (
        'шт', 'м', 'компл', 'комплект', 'упак', 'кор', 'т', 'тн', 'кг', 'л', 'м2', 'м²',
        'пог.м', 'погм', 'п.м', 'пм',
    ):
        return True
    if u_compact.startswith('компл'):
        return True
    if 'пог' in u_compact and 'м' in u_compact:
        return True
    return False


def _pdf_parse_money_cell(cell):
    """Число из ячейки цены/суммы (запятая, пробелы, без «%» для колонки маржи)."""
    if cell is None:
        return None
    s = str(cell).strip().replace('\xa0', '').replace(' ', '')
    if not s or '%' in s:
        return None
    s = s.replace(',', '.')
    try:
        v = float(s)
    except ValueError:
        return None
    if v != v or v < 0 or v > 1_000_000:
        return None
    return v


def _pdf_parse_qty_cell(cell):
    """
    Количество только из ячейки колонки «Кол-во» (не смешивать с ценами/артикулами из других колонок).
    Поддержка «10», «10,5», «10 шт», «10.00 шт.», «2 короб».
    """
    if cell is None:
        return None
    s0 = str(cell).strip().replace('\xa0', ' ')
    if not s0:
        return None
    v = _pdf_parse_money_cell(s0)
    if v is not None and 0.0001 <= v <= 100_000:
        return v
    m = re.search(
        r'(\d+(?:[.,]\d+)?)\s*(?:шт\.?|шт\b|м\b|компл\.?|комплект|пог\.?\s*м|п\.м\.?|пм\b|кор\.?|упак)\b',
        s0,
        re.I,
    )
    if m:
        try:
            v = float(m.group(1).replace(',', '.').replace(' ', ''))
            if 0.0001 <= v <= 100_000:
                return v
        except ValueError:
            pass
    m = re.search(r'(\d+(?:[.,]\d+)?)', s0.replace(',', '.'))
    if m:
        try:
            v = float(m.group(1).replace(',', '.'))
            if 0.0001 <= v <= 100_000:
                return v
        except ValueError:
            pass
    return None


def _pdf_table_tail_purchase_and_sum(
    cells, retail_idx, qty_hint=None, list_price=None, for_wholesale=False
):
    """
    После колонки «розница»: закуп, (НДС 1,2 / маржа %), сумма.
    Не считать первое число закупом, если это коэф. НДС или если другое число даёт сумму×кол-во.
    Для опта и сметы «сумма = розница×кол-во»: закуп — не первая ячейка хвоста (часто маржа % без %),
    а число строго меньше розницы (обычно max из подходящих).
    """
    money_vals = []
    for i in range(retail_idx + 1, len(cells)):
        s = str(cells[i] or '')
        if '%' in s:
            continue
        v = _pdf_parse_money_cell(s)
        if v is not None:
            money_vals.append(v)
    if not money_vals:
        return None, None

    s_last = money_vals[-1]
    head = money_vals[:-1] if len(money_vals) >= 2 else []

    try:
        q = float(qty_hint) if qty_hint is not None else 0.0
    except (TypeError, ValueError):
        q = 0.0

    if for_wholesale and list_price is not None and q > 0 and len(money_vals) >= 2:
        try:
            lp = float(list_price)
            sf = float(s_last)
        except (TypeError, ValueError):
            lp = sf = 0.0
        if (
            lp > 0
            and sf > 0
            and _pdf_line_amount_matches_unit_qty(lp, q, sf)
            and head
        ):
            below = []
            for v in head:
                fv = float(v)
                if 1.17 <= fv <= 1.22:
                    continue
                if fv < lp - 1e-5:
                    below.append(fv)
            if below:
                return (max(below), s_last)

    if q > 0 and len(head) >= 1:
        match_vs_sum = []
        for v in head:
            fv = float(v)
            if 1.17 <= fv <= 1.22:
                continue
            if _pdf_line_amount_matches_unit_qty(fv, q, s_last):
                match_vs_sum.append(fv)
        if len(match_vs_sum) == 1:
            return match_vs_sum[0], s_last
        if len(match_vs_sum) > 1:
            if list_price is not None:
                try:
                    lp = float(list_price)
                    below = [x for x in match_vs_sum if x < lp - 1e-6]
                    if len(below) == 1:
                        return below[0], s_last
                    if below:
                        return min(below), s_last
                except (TypeError, ValueError):
                    pass
            return min(match_vs_sum), s_last

    if len(money_vals) >= 2:
        first_f = float(money_vals[0])
        if 1.17 <= first_f <= 1.22:
            if len(money_vals) >= 3:
                return float(money_vals[1]), s_last
            return None, s_last
        return money_vals[0], s_last
    return None, money_vals[0]


def _pdf_line_amount_matches_unit_qty(unit_price, qty, line_sum, rel_tol=0.025):
    """Проверка: цена_за_ед × кол-во ≈ сумма строки (как в счёте/смете)."""
    try:
        u = float(unit_price)
        q = float(qty)
        s = float(line_sum)
    except (TypeError, ValueError):
        return False
    if u <= 0 or q <= 0 or s <= 0:
        return False
    prod = u * q
    tol = max(0.05, rel_tol * s, rel_tol * prod)
    return abs(prod - s) <= tol


def _pdf_resolve_wholesale_unit_table(row, table_layout):
    """
    Цена закупа за единицу из ячеек PDF (у поставщика обычно без НДС). Перед ответом API для режима
    wholesale умножается на коэффициент НДС. Сумма в смете часто = qty×розница, поэтому не выбираем закуп
    только по совпадению с суммой. Опираемся на колонку закупа из хвоста (уже без «1,2 НДС»)
    и на то, что закуп < розницы в той же строке.
    """
    qty = _pdf_extract_qty(row, from_cart_pdf=False, table_layout=table_layout)
    if qty is None or qty <= 0:
        qty = 1.0
    ri = table_layout['retail_idx']
    list_col = _pdf_parse_money_cell(row[ri]) if ri < len(row) else None
    p_hint = table_layout.get('purchase_from_tail')
    s_line = table_layout.get('line_sum_from_tail')

    if p_hint and 1.17 <= float(p_hint) <= 1.22:
        p_hint = None

    use_hint = float(p_hint) if p_hint is not None and float(p_hint) > 0 else None
    hint_rejected_low = False

    retail_sum_line = (
        s_line is not None
        and list_col is not None
        and qty > 0
        and _pdf_line_amount_matches_unit_qty(list_col, qty, s_line)
    )

    if use_hint is not None and list_col is not None:
        ph, lp = float(use_hint), float(list_col)
        if ph < lp - 1e-6:
            if lp > 40 and ph < lp * 0.05:
                use_hint = None
                hint_rejected_low = True
            else:
                return round(ph, 4)
        elif ph > lp * 1.35 + 0.01:
            trial = ph / float(qty)
            if 0.01 < trial < lp - 1e-6:
                return round(trial, 4)
            if s_line and s_line > 0:
                per = float(s_line) / float(qty)
                if 0.01 < per < lp - 1e-6:
                    return round(per, 4)
            return round(lp, 4)
    elif use_hint is not None:
        return round(float(use_hint), 4)
    elif list_col is not None and not hint_rejected_low and not retail_sum_line:
        return round(float(list_col), 4)

    if s_line and s_line > 0:
        per = float(s_line) / float(qty)
        if list_col is not None and per < float(list_col) - 1e-6:
            return round(per, 4)
        if list_col is None:
            return round(per, 4)

    if retail_sum_line and list_col is not None:
        lc = float(list_col)
        for j in range(ri + 1, len(row)):
            v = _pdf_parse_money_cell(row[j])
            if v is None:
                continue
            fv = float(v)
            if 1.17 <= fv <= 1.22:
                continue
            if fv < lc - 1e-5:
                return round(fv, 4)
    return None


def _pdf_header_norm(s):
    t = str(s or '').lower().replace('ё', 'е').strip()
    t = re.sub(r'[\s\xa0]+', ' ', t)
    t = re.sub(r'[^a-zа-я0-9% ]', '', t)
    return t


def _pdf_row_has_mega_cell(row, max_len=280):
    """pdfplumber иногда склеивает всю страницу в одну ячейку — ломает заголовок."""
    if not row:
        return False
    for cell in row:
        if cell is not None and len(str(cell).strip()) > max_len:
            return True
    return False


def _pdf_filter_table_rows(rows):
    """Убрать «склеенные» строки и пустые."""
    out = []
    for row in rows or []:
        if not row or not any(cell for cell in row if cell):
            continue
        if _pdf_row_has_mega_cell(row):
            continue
        out.append(row)
    return out


def _pdf_header_looks_like_qty(h):
    """Заголовок колонки количества (в т.ч. «кол-\\nчество» из pdfplumber)."""
    if not h:
        return False
    return (
        'колво' in h
        or 'кол во' in h
        or 'колич' in h
        or h.startswith('кол')
        or 'чество' in h
    )


def _pdf_supplier_pricelist_column_sets():
    """Два типичных формата счёта: с пустой col0 (счёт) и без (прайс РРЦ)."""
    return (
        {'num_idx': 1, 'name_idx': 2, 'qty_idx': 3, 'unit_idx': 4, 'price_idx': 5, 'sum_idx': 6},
        {'num_idx': 0, 'name_idx': 1, 'qty_idx': 2, 'unit_idx': 3, 'price_idx': 4, 'sum_idx': 5},
    )


def _pdf_detect_supplier_pricelist_row_layout(row, import_mode='retail'):
    """
    Счёт / прайс РРЦ: № | наименование | кол-во | ед. | цена | сумма.
    Поддерживает ведущую пустую колонку (счёт) и формат без неё (Цены РРЦ).
    """
    if not row or len(row) < 6:
        return None
    cells = [str(c or '').strip() for c in row]
    if _pdf_row_has_mega_cell(cells):
        return None
    for cols in _pdf_supplier_pricelist_column_sets():
        num_idx = cols['num_idx']
        name_idx = cols['name_idx']
        qty_idx = cols['qty_idx']
        unit_idx = cols['unit_idx']
        price_idx = cols['price_idx']
        sum_idx = cols['sum_idx']
        if len(cells) <= price_idx:
            continue
        num_cell = cells[num_idx] if num_idx < len(cells) else ''
        if not re.fullmatch(r'\d{1,4}', num_cell or ''):
            continue
        name_cell = (cells[name_idx] if name_idx < len(cells) else '').strip()
        if len(name_cell) < 3:
            continue
        if not _pdf_cell_is_unit(cells[unit_idx] if unit_idx < len(cells) else ''):
            continue
        qty_hint = _pdf_parse_qty_cell(cells[qty_idx] if qty_idx < len(cells) else '')
        list_price = _pdf_parse_money_cell(cells[price_idx] if price_idx < len(cells) else '')
        if qty_hint is None or list_price is None:
            continue
        purchase_hint, sum_hint = _pdf_table_tail_purchase_and_sum(
            cells,
            price_idx,
            qty_hint,
            list_price,
            for_wholesale=(str(import_mode).lower() == 'wholesale'),
        )
        return {
            'name_idx': name_idx,
            'unit_idx': unit_idx,
            'qty_idx': qty_idx,
            'retail_idx': price_idx,
            'purchase_idx': None,
            'sum_idx': sum_idx if sum_idx < len(cells) else price_idx + 1,
            'purchase_from_tail': purchase_hint,
            'line_sum_from_tail': sum_hint,
            'layout_source': 'supplier_pricelist',
        }
    return None


def _pdf_is_supplier_pricelist_layout(table_layout):
    if not table_layout:
        return False
    if table_layout.get('layout_source') == 'supplier_pricelist':
        return True
    try:
        return (
            int(table_layout.get('name_idx')) == 2
            and int(table_layout.get('qty_idx')) == 3
            and int(table_layout.get('unit_idx')) == 4
            and int(table_layout.get('retail_idx')) == 5
        )
    except (TypeError, ValueError):
        return False


def _pdf_extract_wholesale_unit_supplier(row, table_layout):
    """
    Счёт поставщика (№ 6005): оптовая цена за ед. в колонке «Цена» (индекс 5).
    Колонки НДС/итого справа не используем — там другие суммы (10,68 / 64,06).
    """
    if not row or not _pdf_is_supplier_pricelist_layout(table_layout):
        return None
    try:
        ri = int(table_layout['retail_idx'])
    except (TypeError, ValueError, KeyError):
        return None
    if ri >= len(row):
        return None
    unit = _pdf_parse_money_cell(row[ri])
    if unit is None or unit <= 0:
        return None
    return round(float(unit), 4)


def _pdf_infer_header_layout_from_data(rows, import_mode='retail'):
    """Если строка заголовков не распознана — берём раскладку с первой нормальной строки данных."""
    hits = 0
    sample = None
    for row in rows or []:
        layout = _pdf_detect_supplier_pricelist_row_layout(row, import_mode)
        if layout is None:
            continue
        hits += 1
        if sample is None:
            sample = layout
        if hits >= 2:
            break
    if not sample:
        return None
    return {
        'name_idx': sample['name_idx'],
        'unit_idx': sample['unit_idx'],
        'qty_idx': sample['qty_idx'],
        'retail_idx': sample['retail_idx'],
        'purchase_idx': sample.get('purchase_idx'),
        'header_cells': [
            '№',
            'Наименование',
            'Кол-во',
            'Ед. изм.',
            'Цена' if import_mode != 'wholesale' else 'Цена опт',
            'Стоимость',
        ],
        'sum_idx': 6,
        'layout_source': 'supplier_pricelist',
    }


def _pdf_detect_header_layout(rows, import_mode='retail'):
    """
    Один раз на файл: определить индексы колонок по строке заголовков.
    Это убирает гадание по данным строкам.
    """
    if not rows:
        return None
    rows = _pdf_filter_table_rows(rows)
    limit = min(len(rows), 80)
    for row in rows[:limit]:
        cells = [str(c or '').strip() for c in row]
        if len(cells) < 4 or _pdf_row_has_mega_cell(cells):
            continue
        n = [_pdf_header_norm(c) for c in cells]
        name_idx = unit_idx = qty_idx = retail_idx = purchase_idx = None
        for i, h in enumerate(n):
            if not h:
                continue
            if name_idx is None and (
                'наимен' in h
                or 'товар' in h
                or 'прайс' in h
                or 'материал' in h
                or 'номенклат' in h
            ):
                name_idx = i
            if unit_idx is None and (h.startswith('ед') or 'изм' in h or h == 'ед'):
                unit_idx = i
            if qty_idx is None and _pdf_header_looks_like_qty(h):
                qty_idx = i
            if retail_idx is None and (
                'розн' in h or 'ррц' in h or h == 'цена' or 'ценопт' in h or h == 'ценаопт'
            ):
                retail_idx = i
            if purchase_idx is None and ('закуп' in h or 'опт' in h or 'себестоим' in h):
                purchase_idx = i
            if name_idx is None and h in ('n', 'no', 'пп', 'п/п') and i + 1 < len(cells):
                name_idx = i + 1

        if name_idx is None or qty_idx is None or retail_idx is None:
            continue
        return {
            'name_idx': name_idx,
            'unit_idx': unit_idx,
            'qty_idx': qty_idx,
            'retail_idx': retail_idx,
            'purchase_idx': purchase_idx,
            'header_cells': [str(c or '').strip() for c in cells],
        }
    return _pdf_infer_header_layout_from_data(rows, import_mode)


def _pdf_parse_manual_header_layout(req):
    """
    Ручной выбор колонок от пользователя (1-based индексы из UI).
    Возвращает layout с 0-based индексами или None.
    """
    def _to_idx(name):
        raw = req.form.get(name) if hasattr(req, 'form') else None
        if raw is None:
            raw = req.args.get(name) if hasattr(req, 'args') else None
        if raw is None:
            return None
        s = str(raw).strip()
        if not s:
            return None
        try:
            v = int(s)
        except (TypeError, ValueError):
            return None
        return (v - 1) if v >= 1 else None

    name_idx = _to_idx('col_name')
    qty_idx = _to_idx('col_qty')
    retail_idx = _to_idx('col_retail')
    if name_idx is None or qty_idx is None or retail_idx is None:
        return None
    return {
        'name_idx': name_idx,
        'unit_idx': _to_idx('col_unit'),
        'qty_idx': qty_idx,
        'retail_idx': retail_idx,
        'purchase_idx': _to_idx('col_purchase'),
    }


def _pdf_detect_table_material_row(row, import_mode='retail', header_layout=None):
    """
    Типичная таблица как на скрине сметы: Наименование | Ед. | Кол-во | Розница | Закуп | …
    Либо с ведущим № п/п: № | Наименование | Ед. | Кол-во | Розница | …
    import_mode: для wholesale иначе выбирается колонка закупа при «сумма = розница×qty».
    """
    if not row or len(row) < 4:
        return None
    cells = [str(c or '').strip() for c in row]
    name_idx = unit_idx = qty_idx = retail_idx = None
    purchase_idx = None

    if header_layout:
        name_idx = header_layout.get('name_idx')
        unit_idx = header_layout.get('unit_idx')
        qty_idx = header_layout.get('qty_idx')
        retail_idx = header_layout.get('retail_idx')
        purchase_idx = header_layout.get('purchase_idx')
        required = [name_idx, qty_idx, retail_idx]
        if any(x is None for x in required):
            return None
        if max(int(name_idx), int(qty_idx), int(retail_idx)) >= len(cells):
            return None
    else:
        if len(cells) >= 5 and re.fullmatch(r'\d{1,4}', cells[0] or '') and _pdf_cell_is_unit(cells[2]):
            name_idx, unit_idx = 1, 2
            qty_idx, retail_idx = 3, 4
        elif _pdf_cell_is_unit(cells[1]):
            name_idx, unit_idx = 0, 1
            qty_idx, retail_idx = 2, 3
        else:
            return None

    name_cell = (cells[name_idx] or '').strip()
    if len(name_cell) < 2:
        return None
    low = name_cell.lower()
    if _PDF_HEADER_START.search(low) or _PDF_TOTALS_ROW.search(low):
        return None

    qty_hint = _pdf_parse_qty_cell(cells[qty_idx])
    if qty_hint is None:
        qty_hint = _pdf_parse_money_cell(cells[qty_idx])
    list_price = _pdf_parse_money_cell(cells[retail_idx])
    purchase_hint, sum_hint = _pdf_table_tail_purchase_and_sum(
        cells,
        retail_idx,
        qty_hint,
        list_price,
        for_wholesale=(str(import_mode).lower() == 'wholesale'),
    )
    # Для опта в табличных счетах закуп обычно в соседней колонке сразу после «Розница».
    # Если она валидна — приоритетнее эвристик хвоста (где может быть маржа/служебные числа).
    if str(import_mode).lower() == 'wholesale':
        direct_idx = purchase_idx if purchase_idx is not None else (retail_idx + 1)
        direct_purchase = _pdf_parse_money_cell(cells[direct_idx]) if direct_idx < len(cells) else None
        if direct_purchase is not None and direct_purchase > 0:
            if list_price is None or direct_purchase < (float(list_price) * 1.25 + 0.01):
                purchase_hint = float(direct_purchase)
    return {
        'name_idx': name_idx,
        'unit_idx': unit_idx,
        'qty_idx': qty_idx,
        'retail_idx': retail_idx,
        'purchase_idx': purchase_idx,
        'purchase_from_tail': purchase_hint,
        'line_sum_from_tail': sum_hint,
    }


def _pdf_table_layout_header_fallback(row, import_mode, header_layout):
    """
    Если _pdf_detect_table_material_row вернула None (короткое наименование, «битая» строка и т.д.),
    но по файлу уже есть header_layout — всё равно фиксируем индексы колонок из заголовка.
    Иначе _pdf_extract_qty/_extract_unit уходят в разбор всей строки и смешивают колонки.
    """
    if not row or not header_layout:
        return None
    cells = [str(c or '').strip() for c in row]
    if len(cells) < 4:
        return None
    hl = header_layout
    name_idx = hl.get('name_idx')
    qty_idx = hl.get('qty_idx')
    retail_idx = hl.get('retail_idx')
    if name_idx is None or qty_idx is None or retail_idx is None:
        return None
    try:
        mx = max(int(name_idx), int(qty_idx), int(retail_idx))
    except (TypeError, ValueError):
        return None
    unit_idx = hl.get('unit_idx')
    purchase_idx = hl.get('purchase_idx')
    if unit_idx is not None:
        try:
            mx = max(mx, int(unit_idx))
        except (TypeError, ValueError):
            return None
    if purchase_idx is not None:
        try:
            mx = max(mx, int(purchase_idx))
        except (TypeError, ValueError):
            return None
    if mx >= len(cells):
        return None
    row_join = ' '.join(cells).lower()
    if _PDF_TOTALS_ROW.search(row_join):
        return None

    qty_hint = _pdf_parse_qty_cell(cells[qty_idx])
    if qty_hint is None:
        qty_hint = _pdf_parse_money_cell(cells[qty_idx])
    list_price = _pdf_parse_money_cell(cells[retail_idx])
    purchase_hint, sum_hint = _pdf_table_tail_purchase_and_sum(
        cells,
        retail_idx,
        qty_hint,
        list_price,
        for_wholesale=(str(import_mode).lower() == 'wholesale'),
    )
    if str(import_mode).lower() == 'wholesale':
        direct_idx = purchase_idx if purchase_idx is not None else (retail_idx + 1)
        direct_purchase = _pdf_parse_money_cell(cells[direct_idx]) if direct_idx < len(cells) else None
        if direct_purchase is not None and direct_purchase > 0:
            if list_price is None or direct_purchase < (float(list_price) * 1.25 + 0.01):
                purchase_hint = float(direct_purchase)
    return {
        'name_idx': name_idx,
        'unit_idx': unit_idx,
        'qty_idx': qty_idx,
        'retail_idx': retail_idx,
        'purchase_idx': purchase_idx,
        'purchase_from_tail': purchase_hint,
        'line_sum_from_tail': sum_hint,
    }


def _pdf_adjust_row_layout(row, layout):
    """
    В PDF часто нет ячейки «ед.» в строке данных — тогда цены сдвинуты влево на одну колонку.
    """
    if not layout or not row:
        return layout
    layout = dict(layout)
    cells = [str(c or '').strip() for c in row]
    ui = layout.get('unit_idx')
    if ui is None:
        return layout
    try:
        ui = int(ui)
    except (TypeError, ValueError):
        return layout
    if ui >= len(cells) or not cells[ui]:
        return layout
    if _pdf_cell_is_unit(cells[ui]):
        return layout
    if _pdf_parse_money_cell(cells[ui]) is None:
        return layout
    shifted = False
    for key in ('qty_idx', 'retail_idx', 'purchase_idx'):
        if layout.get(key) is None:
            continue
        try:
            layout[key] = int(layout[key]) - 1
            shifted = True
        except (TypeError, ValueError):
            pass
    if not shifted:
        return layout
    layout['unit_idx'] = None
    try:
        qi = int(layout['qty_idx'])
        ri = int(layout['retail_idx'])
    except (TypeError, ValueError, KeyError):
        return layout
    qty_hint = _pdf_parse_qty_cell(cells[qi]) if qi < len(cells) else None
    if qty_hint is None:
        qty_hint = _pdf_parse_money_cell(cells[qi]) if qi < len(cells) else None
    list_price = _pdf_parse_money_cell(cells[ri]) if ri < len(cells) else None
    purchase_hint, sum_hint = _pdf_table_tail_purchase_and_sum(
        cells,
        ri,
        qty_hint,
        list_price,
        for_wholesale=(str(layout.get('_import_mode') or '') == 'wholesale'),
    )
    layout['purchase_from_tail'] = purchase_hint
    layout['line_sum_from_tail'] = sum_hint
    return layout


def _pdf_table_layout_for_row(row, import_mode, header_layout):
    """Раскладка колонок: полная проверка строки или fallback по header_layout файла."""
    if _pdf_row_has_mega_cell(row):
        return None
    layout = _pdf_detect_supplier_pricelist_row_layout(row, import_mode)
    if layout is None:
        layout = _pdf_detect_table_material_row(row, import_mode, header_layout)
    if layout is None:
        layout = _pdf_table_layout_header_fallback(row, import_mode, header_layout)
    if layout is not None:
        layout['_import_mode'] = import_mode
        layout = _pdf_adjust_row_layout(row, layout)
        layout.pop('_import_mode', None)
    return layout


def _pdf_extract_qty_from_line_sum(table_layout, retail_unit):
    """Кол-во = сумма строки / розница, если в ячейке «кол-во» оказалась цена."""
    if not table_layout or retail_unit is None:
        return None
    s_line = table_layout.get('line_sum_from_tail')
    try:
        s = float(s_line)
        r = float(retail_unit)
    except (TypeError, ValueError):
        return None
    if s <= 0 or r <= 0:
        return None
    q = s / r
    if 0.0001 <= q <= 100_000:
        return round(q, 4)
    return None


def _pdf_finalize_quantity(qty, pdf_retail, pdf_purchase=None, table_layout=None):
    """
    Типичная ошибка PDF: в «кол-во» попала розница (82.75 = 82.75), закуп за единицу нормальный (8.35).
    """
    try:
        q = float(qty)
        r = float(pdf_retail)
        p = float(pdf_purchase) if pdf_purchase is not None else None
    except (TypeError, ValueError):
        return qty
    if r <= 0 or q <= 0:
        return qty
    if abs(q - r) / r > 0.04:
        return qty
    if p is not None and p > 0 and p < r * 0.35:
        return 1.0
    if table_layout:
        s_line = table_layout.get('line_sum_from_tail')
        if s_line is not None:
            try:
                s = float(s_line)
                if abs(s - r) <= max(0.05, 0.02 * abs(s)):
                    return 1.0
            except (TypeError, ValueError):
                pass
        q2 = _pdf_extract_qty_from_line_sum(table_layout, r)
        if q2 is not None and abs(q2 - r) / r > 0.04:
            return q2
    if q >= 8 and abs(q - r) / r <= 0.01:
        return 1.0
    return qty


def _pdf_unit_from_table_cell(cell):
    """Нормализация ед. изм. для сметы из ячейки таблицы."""
    if not cell:
        return 'шт'
    s = str(cell).strip().lower().replace('\xa0', '')
    if 'пог' in s and 'м' in s:
        return 'м'
    if s.startswith('компл'):
        return 'компл'
    if re.match(r'^м\b', s) or s in ('м.', 'м'):
        return 'м'
    return 'шт'


def _pdf_normalize_for_compare(s):
    """Агрессивная нормализация для сравнения с каталогом (разный пунктуация, дроби)."""
    if not s:
        return ''
    t = str(s).lower().replace('ё', 'е')
    t = re.sub(r'[«»"\'`]', ' ', t)
    t = re.sub(r'[\[\]{}]', ' ', t)
    t = re.sub(r'(\d)\s*\.\s*(\d+)\s*/', r'\1 \2/', t)
    t = re.sub(r'(\d)\s*/\s*(\d+)', r'\1/\2', t)
    t = re.sub(r'(\d)\s*х\s*(\d)', r'\1x\2', t, flags=re.I)
    t = re.sub(r'\s*[-–—]{1,3}\s*', ' ', t)
    t = re.sub(r'\.{2,}', ' ', t)
    t = re.sub(r'[^\w\s/+\-°*]', ' ', t, flags=re.UNICODE)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def _pdf_compare_nospace(s):
    """Та же строка без пробелов — «КГ тп» и «КГтп» дают ближе ratio."""
    return re.sub(r'\s+', '', s or '')


def _pdf_normalize_name_for_index(name):
    """Нормализация для индекса каталога: не удаляем все цифры (различаем 600×600 и 300×300)."""
    n = (name or '').lower()
    n = re.sub(r'\([^)]*\)', ' ', n)
    n = re.sub(r'\d+[\.,]\d{2}\b', ' ', n)
    n = re.sub(
        r'\b(китай|россия|испания|италия|германия|турция|польша|евросоюз|беларусь|белоруссия|'
        r'республика\s+польша|пр-во\s*рб)\b',
        ' ',
        n,
        flags=re.I,
    )
    n = re.sub(
        r'\b(в\s*наличии|под\s*заказ|срок|доставка|шт\.?|упак|пог\.?\s*м|пог\.?\s*м\.?)\b',
        ' ',
        n,
        flags=re.I,
    )
    n = re.sub(r'(\d{2,4})\s*[xх]\s*(\d{2,4})', r'\1x\2', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def _pdf_significant_words(text_lower):
    words = re.findall(r'[а-яёa-z0-9][а-яёa-z0-9\-]{1,}', text_lower, flags=re.I)
    out = []
    for w in words:
        wl = w.lower()
        if wl in _PDF_STOPWORDS or len(wl) < 2:
            continue
        if len(wl) == 2 and not re.search(r'\d', wl):
            continue
        out.append(wl)
    return out[:14]


def _pdf_floats_in_order_from_text(s):
    """Числа из строки по порядку (в т.ч. «1 040,18»)."""
    if not s:
        return []
    out = []
    for m in re.finditer(r'(?:\d{1,3}(?:\s+\d{3})+|\d+)(?:[.,]\d+)?', str(s)):
        frag = m.group(0).replace(' ', '').replace(',', '.')
        try:
            v = float(frag)
        except ValueError:
            continue
        if 0.0001 <= v <= 1_000_000:
            out.append(v)
    return out


def _pdf_first_matching_triplet(vals):
    """Тройка (кол-во, цена за ед., сумма) где q×p ≈ t."""
    if len(vals) < 3:
        return None
    for i in range(len(vals) - 2):
        a, b, c = vals[i], vals[i + 1], vals[i + 2]
        if a <= 0 or b <= 0 or c <= 0:
            continue
        prod = a * b
        tol = max(0.015 * abs(c), 0.02 * abs(prod), 0.05)
        if abs(prod - c) <= tol:
            return (a, b, c)
    return None


def _pdf_qty_from_triplet(vals):
    """Первое число из тройки q,p,t где q×p ≈ t."""
    t = _pdf_first_matching_triplet(vals)
    return t[0] if t else None


def _pdf_qty_from_layout_triplet(row, table_layout):
    """Кол-во из чисел между колонкой «кол-во» и «суммой», если ячейка кол-ва = цене."""
    if not table_layout or not row:
        return None
    try:
        qi = int(table_layout['qty_idx'])
        ri = int(table_layout['retail_idx'])
    except (TypeError, ValueError):
        return None
    start = qi
    ui = table_layout.get('unit_idx')
    if ui is not None:
        try:
            start = max(qi, int(ui) + 1)
        except (TypeError, ValueError):
            pass
    vals = []
    for j in range(start, len(row)):
        v = _pdf_parse_money_cell(row[j])
        if v is not None:
            vals.append(v)
    q = _pdf_qty_from_triplet(vals)
    if q is not None:
        return q
    retail = _pdf_parse_money_cell(row[ri]) if ri < len(row) else None
    s_line = table_layout.get('line_sum_from_tail')
    if retail and s_line and float(retail) > 0:
        q = float(s_line) / float(retail)
        if 0.0001 <= q <= 100_000:
            return round(q, 4)
    return None


def _pdf_qty_near_price(qty, *refs, rtol=0.035):
    try:
        qf = float(qty)
    except (TypeError, ValueError):
        return False
    for ref in refs:
        if ref is None:
            continue
        try:
            rf = float(ref)
        except (TypeError, ValueError):
            continue
        if rf > 0 and abs(qf - rf) / rf <= rtol:
            return True
    return False


def _pdf_row_numeric_tail_vals(row):
    """Числа из ячеек строки (кроме первой — часто № или длинное наименование)."""
    ordered = []
    for idx, cell in enumerate(row):
        if idx == 0:
            continue
        s = str(cell or '').strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
        if not s:
            continue
        try:
            v = float(s)
        except ValueError:
            continue
        if 0.05 <= v <= 100000:
            ordered.append((idx, v))
    return [v for _, v in ordered]


def _pdf_retail_price_match(catalog_retail, pdf_price, abs_tol=0.02):
    """Совпадение розничной цены каталога с ценой из PDF (2 знака, допуск по копейкам)."""
    if pdf_price is None:
        return False
    try:
        p = float(pdf_price)
    except (TypeError, ValueError):
        return False
    if p <= 0:
        return False
    try:
        c = float(catalog_retail) if catalog_retail is not None else None
    except (TypeError, ValueError):
        c = None
    if c is None or c <= 0:
        return False
    return abs(round(c, 2) - round(p, 2)) <= abs_tol


def _pdf_purchase_for_file_retail(file_retail, cat_item, qty=1.0):
    """Закупка для строки сметы при рознице из PDF: сохраняем отношение закуп/розница из каталога."""
    try:
        fr = float(file_retail)
        cr = float(cat_item.get('retail_price') or 0)
        cp = float(cat_item.get('purchase_price') or 0)
        q = float(qty or 1)
    except (TypeError, ValueError):
        return None
    if fr <= 0:
        return None
    cp = _normalize_unit_purchase_price(cp, cr, qty=q)
    if cr > 0 and cp >= 0:
        pu = round(fr * (cp / cr), 4)
        return _normalize_unit_purchase_price(pu, fr, qty=q)
    return round(fr * 0.7, 2)


def _pdf_try_cart_row_qty_unit_price(row, from_cart_pdf=False):
    """
    Только для строк, собранных из PDF корзины (akvabreg и т.п.): [наименование, цена_за_ед, кол-во, сумма, артикул].
    Для старых табличных PDF (ИТП и др.) не вызывать — там порядок колонок другой, работает тройка q×p≈s.
    """
    if not from_cart_pdf:
        return None, None
    if not row or len(row) != 5:
        return None, None
    name = str(row[0] or '').strip()
    if len(name) < 5:
        return None, None
    try:
        unit_p = float(str(row[1]).replace(',', '.').replace(' ', ''))
        qty = float(str(row[2]).replace(',', '.').replace(' ', ''))
        total = float(str(row[3]).replace(',', '.').replace(' ', ''))
    except (ValueError, TypeError):
        return None, None
    if unit_p <= 0 or qty <= 0 or total <= 0:
        return None, None
    tol = max(0.02, 0.015 * abs(total), 0.02 * abs(unit_p * qty))
    if abs(unit_p * qty - total) > tol:
        return None, None
    art = str(row[4] or '').strip().replace(' ', '')
    if art and not re.fullmatch(r'\d{8,14}', art):
        return None, None
    return qty, unit_p


def _pdf_extract_pdf_retail_unit_price(row, from_cart_pdf=False, table_layout=None):
    """Розничная цена за единицу из строки прайса (колонка «цена», не сумма)."""
    _cq, cpu = _pdf_try_cart_row_qty_unit_price(row, from_cart_pdf)
    if cpu is not None:
        return cpu
    if table_layout is not None:
        ri = table_layout['retail_idx']
        if ri < len(row):
            v = _pdf_parse_money_cell(row[ri])
            if v is not None and v > 0:
                return v
    whole = ' '.join(str(c or '') for c in row)
    vals = _pdf_row_numeric_tail_vals(row)
    t = _pdf_first_matching_triplet(vals)
    if t:
        return t[1]
    t = _pdf_first_matching_triplet(_pdf_floats_in_order_from_text(whole))
    if t:
        return t[1]
    if len(vals) == 2:
        v0, v1 = vals[0], vals[1]
        if 0.01 <= v1 <= 1_000_000:
            return v1
    return None


def _pdf_extract_pdf_unit_price_with_vat(row, from_cart_pdf=False, table_layout=None):
    """
    Цена закупа за единицу из оптового PDF (как в документе; чаще без НДС).
    Для cart-ready PDF возвращаем розничную цену за единицу (режим корзины).
    Для счёта поставщика (supplier_pricelist) — цена из колонки «Цена», НДС по сумме строки.
    """
    if from_cart_pdf:
        return _pdf_extract_pdf_retail_unit_price(row, from_cart_pdf=True, table_layout=None)
    if _pdf_is_supplier_pricelist_layout(table_layout):
        w = _pdf_extract_wholesale_unit_supplier(row, table_layout)
        if w is not None:
            return w
    if table_layout is not None:
        pidx = table_layout.get('purchase_idx')
        if pidx is not None and pidx < len(row):
            pv = _pdf_parse_money_cell(row[pidx])
            if pv is not None and pv > 0:
                return round(float(pv), 4)
        w = _pdf_resolve_wholesale_unit_table(row, table_layout)
        if w is not None:
            return w
        ri = table_layout['retail_idx']
        list_top = _pdf_parse_money_cell(row[ri]) if ri < len(row) else None
        for j in range(ri + 1, min(ri + 5, len(row))):
            v = _pdf_parse_money_cell(row[j])
            if v is None:
                continue
            fv = float(v)
            if 1.17 <= fv <= 1.22:
                continue
            if list_top is not None and fv < float(list_top) - 1e-6:
                return round(fv, 4)
    qty = _pdf_extract_qty(row, from_cart_pdf=False, table_layout=table_layout)
    if qty is None or qty <= 0:
        return None
    whole = ' '.join(str(c or '') for c in row)
    vals = _pdf_floats_in_order_from_text(whole)
    money_vals = [v for v in vals if v > 0 and (abs(v - round(v)) > 1e-9 or v >= 1000)]
    if not money_vals:
        return None
    fq = float(qty)
    for v in reversed(money_vals):
        if 1.17 <= v <= 1.22:
            continue
        unit_price = v / fq
        if 0.0001 <= unit_price <= 1_000_000:
            return round(unit_price, 4)
    return None


def _pdf_extract_unit(row, table_layout=None):
    """Ед. изм. из строки PDF (для сметы), не из карточки каталога."""
    if table_layout is not None:
        ui = table_layout.get('unit_idx')
        if ui is not None and int(ui) < len(row):
            return _pdf_unit_from_table_cell(row[ui])
        # Таблица с известными индексами, но колонка «ед.» не найдена в заголовке —
        # не тянуть «м/компл» из наименования или цен (ложные совпадения).
        return 'шт'
    whole = ' '.join(str(c or '') for c in row)
    if re.search(r'\d+(?:[.,]\d+)?\s+пог\.?\s*м\b', whole, re.I):
        return 'м'
    if re.search(r'\d+(?:[.,]\d+)?\s+компл', whole, re.I):
        return 'компл'
    if re.search(
        r'(\d+(?:[.,]\d+)?)\s+м\s+(?=(?:\d{1,3}(?:\s+\d{3})+|\d+)(?:[.,]\d+))',
        whole,
        re.I,
    ):
        return 'м'
    return 'шт'


def _pdf_extract_qty(row, from_cart_pdf=False, table_layout=None):
    """Количество: явные единицы, «N м» перед ценой, тройка qty×price≈total, иначе ячейки."""
    cq, _cpu = _pdf_try_cart_row_qty_unit_price(row, from_cart_pdf)
    if cq is not None:
        return cq
    if table_layout is not None:
        qi = table_layout.get('qty_idx')
        ri = table_layout.get('retail_idx')
        if qi is not None and int(qi) < len(row):
            v = _pdf_parse_qty_cell(row[qi])
            if v is None:
                v = _pdf_parse_money_cell(row[qi])
            retail_v = (
                _pdf_parse_money_cell(row[int(ri)])
                if ri is not None and int(ri) < len(row)
                else None
            )
            if v is not None and 0.0001 <= v <= 100_000:
                if _pdf_qty_near_price(v, retail_v):
                    alt = (
                        _pdf_extract_qty_from_line_sum(table_layout, retail_v)
                        or _pdf_qty_from_layout_triplet(row, table_layout)
                    )
                    v = alt if alt is not None else 1.0
                return v
            # Колонка «кол-во» задана таблицей — не подмешивать числа из цены/наименования.
            alt = _pdf_qty_from_layout_triplet(row, table_layout)
            return alt if alt is not None else 1.0
    whole = ' '.join(str(c or '') for c in row)

    # Смета/счёт: колонка «Кол-во» идёт ПОСЛЕ «шт/м» — «… шт 1 16.36», иначе «25 шт» из «25x25x25 шт» даёт ложное qty.
    for pat in (
        r'(?:шт\.?|шт)\s+(\d+(?:[.,]\d+)?)\b',
        r'(?:компл\.?|комплект)\s+(\d+(?:[.,]\d+)?)\b',
        r'(?:пог\.?\s*м|п\.м\.?)\s+(\d+(?:[.,]\d+)?)\b',
    ):
        m = None
        for m in re.finditer(pat, whole, re.I):
            pass
        if m:
            try:
                v = float(m.group(1).replace(',', '.').replace(' ', ''))
                if 0.01 <= v <= 100_000:
                    return v
            except ValueError:
                pass

    for pat in (
        r'(\d+(?:[.,]\d+)?)\s*(?:шт\.?|шт)\b',
        r'(\d+(?:[.,]\d+)?)\s+пог\.?\s*м\b',
        r'(\d+(?:[.,]\d+)?)\s+компл\.?\b',
    ):
        m = re.search(pat, whole, re.I)
        if m:
            try:
                v = float(m.group(1).replace(',', '.').replace(' ', ''))
                if 0.01 <= v <= 100_000:
                    return v
            except ValueError:
                pass

    m = re.search(
        r'(\d+(?:[.,]\d+)?)\s+м\s+(?=(?:\d{1,3}(?:\s+\d{3})+|\d+)(?:[.,]\d+))',
        whole,
        re.I,
    )
    if m:
        try:
            v = float(m.group(1).replace(',', '.').replace(' ', ''))
            if 0.01 <= v <= 100_000:
                return v
        except ValueError:
            pass

    vals = _pdf_row_numeric_tail_vals(row)
    q3 = _pdf_qty_from_triplet(vals)
    if q3 is not None:
        return q3

    if len(vals) >= 3:
        q3w = _pdf_qty_from_triplet(_pdf_floats_in_order_from_text(whole))
        if q3w is not None:
            return q3w
        q3 = _pdf_qty_from_triplet(vals)
        if q3 is not None:
            return q3
        s_max = max(vals)
        candidates = [x for x in vals if x < s_max * 0.95 and x <= 5000]
        if candidates:
            return min(candidates)
        return 1.0

    if not vals:
        q3w = _pdf_qty_from_triplet(_pdf_floats_in_order_from_text(whole))
        if q3w is not None:
            return q3w
        return 1.0

    if len(vals) == 1:
        return vals[0]

    v0, v1 = vals[0], vals[1]
    if abs(v0 - v1) / max(v0, v1, 1e-9) < 0.03:
        return 1.0
    if v1 > v0 * 50:
        return v0
    if v0 > v1 * 50:
        return v1
    q3w = _pdf_qty_from_triplet(_pdf_floats_in_order_from_text(whole))
    if q3w is not None:
        return q3w
    return min(v0, v1)


def _pdf_sanitize_qty(
    qty, retail_price, purchase_price=0, file_retail=None, file_wholesale=None
):
    """Не подставлять цену/РРЦ в количество (иначе сумма = миллионы)."""
    try:
        q = float(qty)
        rp = float(retail_price or 0)
        pp = float(purchase_price or 0)
        fr = float(file_retail or 0)
        fw = float(file_wholesale or 0)
    except (TypeError, ValueError):
        return 1.0
    if q <= 0 or q != q:
        return 1.0
    if q > 5000:
        return 1.0
    if rp > 0 and q * rp > 400_000:
        return 1.0
    for ref in (rp, pp, fr, fw):
        if ref > 0 and abs(q - ref) / ref <= 0.04:
            return 1.0
    return q


def _pdf_catalog_article_map(catalog_items):
    """Индекс артикулов: как в каталоге, без пробелов, только цифры, без ведущих нулей, хвост EAN-8."""
    m = {}
    for item in catalog_items:
        raw = (item.get('article') or '').strip()
        if not raw:
            continue
        variants = set()
        for base in (raw, raw.upper(), raw.replace(' ', ''), raw.upper().replace(' ', '')):
            if base:
                variants.add(base)
        digits = re.sub(r'\D', '', raw)
        if digits:
            variants.add(digits)
            stripped = digits.lstrip('0')
            if stripped and stripped != digits:
                variants.add(stripped)
            if len(digits) >= 8:
                variants.add(digits[-8:])
        for v in variants:
            if v and 2 <= len(str(v)) <= 40:
                m[str(v)] = item
    return m


def _pdf_lookup_article_in_map(code, catalog_by_article):
    """Поиск позиции по коду с теми же вариантами, что и при индексации."""
    if not code:
        return None
    c0 = str(code).strip()
    candidates = [c0, c0.upper(), c0.replace(' ', ''), c0.upper().replace(' ', '')]
    digits = re.sub(r'\D', '', c0)
    if digits:
        candidates.append(digits)
        stripped = digits.lstrip('0')
        if stripped and stripped != digits:
            candidates.append(stripped)
        if len(digits) >= 8:
            candidates.append(digits[-8:])
    for c in candidates:
        if c and c in catalog_by_article:
            return catalog_by_article[c]
    return None


def _pdf_match_by_article(row_text, row, catalog_by_article):
    """Совпадение по артикулу: скобки, ячейка 8–14 цифр (EAN), «Артикул» в тексте."""
    seen = []
    m = re.search(r'\(([\dA-Za-zА-Яа-яЁё\-/.]{2,})\)', row_text)
    if m:
        seen.append(m.group(1).strip().upper().strip('.'))
    for cell in row:
        t = str(cell).strip()
        td = re.sub(r'\s', '', t)
        if re.fullmatch(r'\d{8,14}', td):
            seen.append(td)
        t_up = td.upper() if td else ''
        if 2 <= len(t_up) <= 36 and re.fullmatch(r'[\dA-ZА-ЯЁ\-/.]+', t_up) and not re.fullmatch(r'\d{1,7}', t_up):
            if t_up not in seen:
                seen.append(t_up)
    for m2 in re.finditer(r'Артикул[:\s]*(\d{8,14})\b', row_text, re.I):
        seen.append(m2.group(1))
    for m2 in re.finditer(r'\b(\d{12,14})\b', row_text):
        g = m2.group(1)
        if g not in seen:
            seen.append(g)

    for code in seen:
        hit = _pdf_lookup_article_in_map(code, catalog_by_article)
        if hit:
            return hit
    return None


def _pdf_match_confidence_display(base_score, lex_ratio, word_coverage):
    """
    Процент для UI: сочетает внутренний score отбора и реальное сходство строк.
    При верных колонках PDF lex_ratio высокий — пользователь снова видит ~90–100%, как раньше.
    """
    try:
        b = float(base_score)
        lr = float(lex_ratio)
        wc = float(word_coverage)
    except (TypeError, ValueError):
        return 0.0
    wc = max(wc, lr * 0.85)
    blend = 0.30 * b + 0.48 * lr + 0.22 * min(1.0, wc)
    return min(100.0, round(blend * 100, 1))


def _pdf_fuzzy_best_match(clean_name, catalog_search_index, pdf_retail_unit_price=None):
    """
    Сопоставление с каталогом. Если задана pdf_retail_unit_price — только позиции с той же розницей,
    внутри них — частичное совпадение по названию (пороги чуть мягче, т.к. цена уже отсекла лишнее).
    """
    from difflib import SequenceMatcher

    if not catalog_search_index:
        return None, 0.0, 0.0, 0.0, 0.0

    price_slice = catalog_search_index
    price_locked = False
    if pdf_retail_unit_price is not None:
        price_slice = [
            tup
            for tup in catalog_search_index
            if _pdf_retail_price_match(tup[0].get('retail_price'), pdf_retail_unit_price)
        ]
        if not price_slice:
            return None, 0.0, 0.0, 0.0, 0.0
        price_locked = True
        catalog_search_index = price_slice

    clean_lower = (clean_name or '').lower()
    clean_cmp = _pdf_normalize_for_compare(clean_name)
    clean_ns = _pdf_compare_nospace(clean_cmp)
    key_words = _pdf_significant_words(clean_lower)
    scored = []

    for item, norm, name_lower, norm_cmp in catalog_search_index:
        r1 = SequenceMatcher(None, clean_lower, name_lower).ratio()
        r2 = SequenceMatcher(None, clean_lower, norm).ratio()
        r3 = SequenceMatcher(None, clean_cmp, norm_cmp).ratio() if norm_cmp else 0.0
        r4 = (
            SequenceMatcher(None, clean_ns, _pdf_compare_nospace(norm_cmp)).ratio()
            if norm_cmp
            else 0.0
        )
        ratio = max(r1, r2, r3, r4)

        nw = len(key_words)
        match_words = sum(
            1
            for kw in key_words
            if kw in name_lower
            or kw in norm
            or (norm_cmp and kw in norm_cmp)
            or (norm_cmp and kw in _pdf_compare_nospace(norm_cmp))
        )
        word_part = (match_words / float(nw)) if nw else ratio

        base = 0.40 * ratio + 0.34 * word_part + 0.14 * r4
        if item.get('brand') and str(item['brand']).lower() in clean_lower:
            base += 0.11
        if item.get('category') and str(item['category']).lower() in clean_lower:
            base += 0.05
        if clean_lower in name_lower or (norm_cmp and clean_cmp in norm_cmp):
            base += 0.09
        elif name_lower in clean_lower or (norm_cmp and norm_cmp in clean_cmp):
            base += 0.06
        elif clean_ns and norm_cmp and clean_ns in _pdf_compare_nospace(norm_cmp):
            base += 0.07

        base = min(base, 0.995)
        scored.append((base, item, ratio, word_part))

    scored.sort(key=lambda x: (-x[0], -x[2]))
    top_s, top_item, top_lex, top_wpart = scored[0]
    second_s = scored[1][0] if len(scored) > 1 else 0.0
    gap = top_s - second_s

    if price_locked:
        if top_s >= 0.34 and gap >= 0.04:
            return top_item, top_s, second_s, top_lex, top_wpart
        if top_s >= 0.22 and gap >= 0.055:
            return top_item, top_s, second_s, top_lex, top_wpart
        if top_s >= 0.18 and top_lex >= 0.30 and gap >= 0.07:
            return top_item, top_s, second_s, top_lex, top_wpart
        return None, top_s, second_s, top_lex, top_wpart

    if top_s >= 0.40:
        return top_item, top_s, second_s, top_lex, top_wpart
    if top_s >= 0.26 and gap >= 0.055:
        return top_item, top_s, second_s, top_lex, top_wpart
    if top_s >= 0.22 and gap >= 0.095:
        return top_item, top_s, second_s, top_lex, top_wpart
    return None, top_s, second_s, top_lex, top_wpart


estimate_bp = Blueprint('estimate', __name__, url_prefix='/estimate')

# CSRF: session['_csrf_token'] выставляет основное приложение (inject_csrf / _csrf_token)

def _require_csrf(f):
    """Декоратор CSRF для estimate модуля"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        if not token or token != session.get('_csrf_token'):
            return jsonify({'error': 'CSRF token missing or invalid'}), 403
        return f(*args, **kwargs)
    return decorated_function

def _require_json():
    """Проверка наличия JSON тела запроса"""
    data = request.get_json(silent=True)
    if not data:
        return None, (jsonify({"error": "Требуется JSON в теле запроса"}), 400)
    return data, None

def _safe_float(value, default=0.0):
    """Безопасное преобразование в float"""
    try:
        return float(value) if value is not None else default
    except (ValueError, TypeError):
        return default


_QTY_GUESS_FOR_LINE_PURCHASE = (
    2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 20, 25, 30, 40, 50, 100,
)


def _normalize_unit_purchase_price(purchase, retail_unit, qty=1.0):
    """
    Закуп в каталоге/смете — всегда за единицу. Старый импорт PDF иногда писал сумму строки
    (закуп×кол-во) в purchase_price → маржа уходила в минус сотнями процентов.
    """
    try:
        p = float(purchase or 0)
        r = float(retail_unit or 0)
        q = float(qty or 1)
    except (TypeError, ValueError):
        return 0.0
    if p <= 0 or r <= 0:
        return round(p, 4) if p > 0 else 0.0
    if p <= r * 1.05:
        return round(p, 4)

    def _try_qty(q_try):
        if q_try <= 1:
            return None
        unit_cand = p / q_try
        if unit_cand <= 0 or unit_cand > r * 1.05:
            return None
        if abs(p - unit_cand * q_try) <= max(0.05, 0.02 * p):
            return round(unit_cand, 4)
        return None

    if q > 1:
        fixed = _try_qty(q)
        if fixed is not None:
            return fixed
    for q_try in _QTY_GUESS_FOR_LINE_PURCHASE:
        fixed = _try_qty(q_try)
        if fixed is not None:
            return fixed
    return round(p, 4)


_ESTIMATE_UNITS = frozenset({'шт', 'м', 'м²', 'м³', 'кг', 'т', 'л', 'компл', 'усл', 'ч'})


def _estimate_match_material_items(estimate_items, pdf_name, retail_price=None):
    """Строки сметы, соответствующие позиции из PDF (имя / фрагмент / розница)."""
    keys = []
    if pdf_name:
        keys.append(_pdf_normalize_name_for_index(pdf_name))
    keys = [k for k in keys if k]
    if not keys and retail_price is None:
        return []
    try:
        fr = float(retail_price) if retail_price is not None else None
    except (TypeError, ValueError):
        fr = None
    out = []
    seen = set()
    for it in estimate_items or []:
        if (it.get('section') or 'material') != 'material':
            continue
        iid = it.get('id')
        if iid in seen:
            continue
        n = _pdf_normalize_name_for_index(it.get('name') or '')
        hit = False
        for k in keys:
            if not k:
                continue
            if n == k or (len(k) >= 8 and k in n) or (len(n) >= 8 and n in k):
                hit = True
                break
        if not hit and fr is not None and fr > 0:
            try:
                pr = float(it.get('price') or 0)
            except (TypeError, ValueError):
                pr = 0
            if pr > 0 and abs(fr - pr) <= max(0.05, 0.01 * pr):
                hit = True
        if hit:
            seen.add(iid)
            out.append(it)
    return out


def _repair_estimate_item_quantity(it):
    """
    Старый импорт PDF иногда писал в quantity ту же цифру, что в price.
    Восстанавливаем по total (qty = total / price) или ставим 1, если total ≈ price.
    """
    if (it.get('section') or 'material') != 'material':
        return False
    q = _safe_float(it.get('quantity'), default=1.0)
    p = _safe_float(it.get('price'))
    t = _safe_float(it.get('total'))
    pu = _safe_float(it.get('purchase_price'))
    if p <= 0 or q <= 0:
        return False
    if abs(q - p) / max(p, 1e-9) > 0.04:
        return False
    new_q = q
    if pu > 0 and pu < p * 0.35:
        new_q = 1.0
    elif t > 0:
        implied = round(t / p, 4)
        if abs(implied - q) / max(q, 1e-9) > 0.12:
            if 0.0001 <= implied <= 5000:
                new_q = implied
            elif abs(t - p) <= max(0.05, 0.02 * abs(t)):
                new_q = 1.0
        elif abs(t - p) <= max(0.05, 0.02 * abs(t)):
            new_q = 1.0
    elif q >= 8:
        new_q = 1.0
    else:
        return False
    if abs(new_q - q) < 1e-6:
        return False
    it['quantity'] = new_q
    pu = _normalize_unit_purchase_price(_safe_float(it.get('purchase_price')), p, qty=new_q)
    it['purchase_price'] = pu
    it['total'] = round(p * new_q, 2)
    it['material_profit'] = round((p - pu) * new_q, 2)
    return True


def _normalize_estimate_unit(unit, section='material'):
    """Ед. изм. для сметы: не допускаем цены и прочие числа вместо «шт»/«м»."""
    default = 'усл' if (section or 'material') == 'work' else 'шт'
    if unit is None:
        return default
    s = str(unit).strip()
    if not s:
        return default
    s_compact = re.sub(r'\s+', '', s.lower().replace('\xa0', '')).rstrip('.')
    if s_compact in ('m', 'пм', 'п.м', 'погм', 'пог.м'):
        return 'м'
    if s_compact.startswith('компл') or s_compact == 'комплект':
        return 'компл'
    if s in _ESTIMATE_UNITS or s_compact in _ESTIMATE_UNITS:
        return s if s in _ESTIMATE_UNITS else s_compact
    try:
        float(s.replace(',', '.').replace(' ', ''))
        return default
    except ValueError:
        pass
    if re.fullmatch(r'[\d.,\s]+', s):
        return default
    if len(s) <= 16 and re.search(r'[a-zA-Zа-яА-Я²³]', s) and not re.fullmatch(r'[\d.,]+', s):
        return s[:20]
    return default


def _norm_item_name(s):
    return re.sub(r"\s+", " ", (s or "").strip()).lower()


def _backfill_estimate_wholesale_from_catalog(user_id, est_id, items):
    """Подставить опт из каталога по имени: пусто/0, или опт совпал с розницей (старый импорт), а в каталоге уже другое."""
    if not items:
        return
    mats = fetch_all(
        "SELECT name, wholesale_price FROM catalog_materials WHERE user_id = ?",
        (user_id,),
    )
    by_name = {}
    for m in mats:
        key = _norm_item_name(m.get("name"))
        if key:
            by_name[key] = _safe_float(m.get("wholesale_price"))
    eps = 0.02
    for it in items:
        if it.get("section") != "material":
            continue
        key = _norm_item_name(it.get("name"))
        if not key:
            continue
        cw = by_name.get(key)
        if cw is None or cw <= 0:
            continue
        line_w = _safe_float(it.get("wholesale_price"))
        line_p = _safe_float(it.get("price"))
        # Опт явно отличается от розницы — не перезаписываем (ручной ввод или уже из opt2)
        if line_w > 0 and abs(line_w - line_p) >= eps:
            continue
        if abs(cw - line_w) < eps:
            continue
        iid = it.get("id")
        if iid:
            execute(
                """UPDATE estimate_items SET wholesale_price = ?
                   WHERE id = ? AND estimate_id = ?
                   AND EXISTS (SELECT 1 FROM estimates WHERE id = ? AND user_id = ?)""",
                (cw, iid, est_id, est_id, user_id),
            )
        it["wholesale_price"] = cw


# ============================
# HTML СТРАНИЦЫ
# ============================

@estimate_bp.route('/')
@login_required
def estimate_list(): return render_template('estimate/list.html')

@estimate_bp.route('/catalog')
@login_required
def catalog_page(): return render_template('estimate/catalog.html')

@estimate_bp.route('/new')
@login_required
def new_estimate(): return render_template('estimate/editor.html', estimate_id=None)

@estimate_bp.route('/<int:estimate_id>')
@login_required
def edit_estimate(estimate_id): return render_template('estimate/editor.html', estimate_id=estimate_id)

# ============================
# API: КАТЕГОРИИ
# ============================

@estimate_bp.route('/api/catalog/categories', methods=['GET'])
@login_required
def api_get_categories():
    cat_type = request.args.get('type', '')  # 'material' или 'work'
    if cat_type:
        return jsonify(fetch_all("SELECT * FROM categories WHERE user_id = ? AND category_type = ? ORDER BY name", (current_user.id, cat_type)))
    return jsonify(fetch_all("SELECT * FROM categories WHERE user_id = ? ORDER BY name", (current_user.id,)))

def _build_db_category_tree(user_id, cat_type):
    """Дерево из таблицы categories (категория → подкатегория) со счётчиками материалов/работ."""
    rows = fetch_all(
        "SELECT id, name, parent_id FROM categories WHERE user_id = ? AND category_type = ? ORDER BY name",
        (user_id, cat_type),
    )
    if not rows:
        return None

    table = 'catalog_materials' if cat_type == 'material' else 'catalog_works'
    items = fetch_all(f"SELECT category FROM {table} WHERE user_id = ?", (user_id,))
    counts = {}
    for it in items:
        k = (it.get('category') or '').strip() or 'Без категории'
        counts[k] = counts.get(k, 0) + 1

    nodes = {}
    for r in rows:
        nodes[r['id']] = {
            'id': r['id'],
            'name': r['name'],
            'parent_id': r.get('parent_id'),
            'children': [],
            'all_descendants': [],
            'count': 0,
            'is_brand_leaf': False,
            'is_db_category': True,
            'filter_level': '0',
        }
    roots = []
    for r in rows:
        n = nodes[r['id']]
        pid = r.get('parent_id')
        if pid and pid in nodes:
            n['filter_level'] = 'sub'
            nodes[pid]['children'].append(n)
        else:
            roots.append(n)

    def walk(node):
        direct = counts.get(node['name'], 0)
        child_names = []
        total = direct
        for ch in node['children']:
            sub_names, sub_total = walk(ch)
            child_names.extend(sub_names)
            total += sub_total
        node['count'] = total
        node['all_descendants'] = child_names
        return [node['name']] + child_names, total

    for root in roots:
        walk(root)
    return roots


def _build_material_category_tree(user_id):
    """Дерево из полей материалов: Категория → Тип → Бренд."""
    items = fetch_all(
        "SELECT category, item_type, brand, name, id FROM catalog_materials WHERE user_id = ?",
        (user_id,),
    )
    tree_map = {}
    for item in items:
        cat = item['category'] or 'Без категории'
        itype = item['item_type'] or 'Другое'
        brand = item['brand'] or 'Без бренда'
        if cat not in tree_map:
            tree_map[cat] = {}
        if itype not in tree_map[cat]:
            tree_map[cat][itype] = {}
        if brand not in tree_map[cat][itype]:
            tree_map[cat][itype][brand] = []
        tree_map[cat][itype][brand].append(item)

    tree = []
    for cat_name in sorted(tree_map.keys()):
        type_branches = tree_map[cat_name]
        type_nodes = []
        cat_total = 0
        skip_type_level = False
        if len(type_branches) == 1:
            only_type = list(type_branches.keys())[0]
            norm_cat = cat_name.lower().replace(' ', '').replace('-', '')
            norm_type = only_type.lower().replace(' ', '').replace('-', '')
            if norm_cat == norm_type or norm_type in norm_cat or norm_cat in norm_type:
                skip_type_level = True

        for type_name in sorted(type_branches.keys()):
            brand_branches = type_branches[type_name]
            brand_nodes = []
            type_total = 0
            for brand_name in sorted(brand_branches.keys()):
                item_list = brand_branches[brand_name]
                type_total += len(item_list)
                brand_nodes.append({
                    'id': None,
                    'name': brand_name,
                    'count': len(item_list),
                    'children': [],
                    'all_descendants': [],
                    'is_brand_leaf': True,
                })
            cat_total += type_total
            type_nodes.append({
                'id': None,
                'name': type_name,
                'count': type_total,
                'children': brand_nodes,
                'all_descendants': [b['name'] for b in brand_nodes],
            })

        if skip_type_level and type_nodes:
            all_brands = type_nodes[0]['children']
            final_children = all_brands
            all_descs = [b['name'] for b in all_brands]
        else:
            final_children = type_nodes
            all_descs = []
            for tn in type_nodes:
                all_descs.append(tn['name'])
                all_descs.extend(tn['all_descendants'])

        tree.append({
            'id': None,
            'name': cat_name,
            'count': cat_total,
            'children': final_children,
            'all_descendants': all_descs,
            'skip_type_level': skip_type_level,
        })
    return tree


@estimate_bp.route('/api/catalog/categories/tree', methods=['GET'])
@login_required
def api_get_categories_tree():
    """Древовидная структура: из БД (категория → подкатегория) или из полей материалов."""
    cat_type = request.args.get('type', 'material')
    db_tree = _build_db_category_tree(current_user.id, cat_type)
    if db_tree is not None:
        return jsonify(db_tree)
    if cat_type != 'material':
        cats = fetch_all(
            "SELECT * FROM categories WHERE user_id = ? AND category_type = ? ORDER BY name",
            (current_user.id, cat_type),
        )
        return jsonify([
            {'id': c['id'], 'name': c['name'], 'children': [], 'all_descendants': [], 'count': 0}
            for c in cats
        ])
    return jsonify(_build_material_category_tree(current_user.id))

@estimate_bp.route('/api/catalog/categories', methods=['POST'])
@login_required
@_require_csrf
def api_add_category():
    data, err = _require_json()
    if err: return err
    try:
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({"error": "Укажите название категории"}), 400
        parent_id = data.get('parent_id')
        if not parent_id or parent_id == 'null':
            parent_id = None
        else:
            parent_id = int(parent_id)
        cat_type = data.get('type', 'material')
        if parent_id:
            parent = fetch_one(
                "SELECT id FROM categories WHERE id = ? AND user_id = ? AND category_type = ?",
                (parent_id, current_user.id, cat_type),
            )
            if not parent:
                return jsonify({"error": "Родительская категория не найдена"}), 400
        new_id = execute(
            "INSERT INTO categories (user_id, name, parent_id, category_type) VALUES (?, ?, ?, ?)",
            (current_user.id, name, parent_id, cat_type),
            return_id=True,
        )
        return jsonify({"status": "ok", "id": new_id}), 201
    except Exception as e:
        logger.error(f"Add category error: {e}")
        return jsonify({"error": "Ошибка при создании категории"}), 400

@estimate_bp.route('/api/parse-pdf', methods=['POST'])
@login_required
def api_parse_pdf():
    """Загрузить PDF, извлечь таблицы, сопоставить с каталогом"""
    try:
        import pdfplumber
        from difflib import SequenceMatcher
        import io, re

        if 'file' not in request.files:
            return jsonify({"error": "Нет файла"}), 400

        file = request.files['file']
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({"error": "Только PDF файлы"}), 400

        # Извлекаем текст и таблицы из PDF
        pdf_file = pdfplumber.open(io.BytesIO(file.read()))

        full_text_chunks = []
        all_rows = []
        for page in pdf_file.pages:
            text = page.extract_text()
            if text:
                full_text_chunks.append(text)
            # Пробуем извлечь таблицы
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if row and any(cell for cell in row if cell):
                            all_rows.append([str(c).strip() if c else '' for c in row])
            else:
                # Если таблиц нет — извлекаем текст построчно
                if text:
                    for line in text.split('\n'):
                        line = line.strip()
                        if line:
                            # Разбиваем по табуляции или нескольким пробелам
                            parts = re.split(r'\t|\s{3,}', line)
                            all_rows.append([p.strip() for p in parts if p.strip()])

        pdf_file.close()

        all_rows = _pdf_filter_table_rows(all_rows)

        mode_raw = (request.form.get('mode') or request.args.get('mode') or 'retail').strip().lower()
        import_mode = 'wholesale' if mode_raw in ('wholesale', 'opt', 'purchase', 'cost') else 'retail'

        full_text = '\n'.join(full_text_chunks)
        table_rows_snapshot = list(all_rows)
        cart_rows = _pdf_parse_cart_ready_format(full_text)
        substantial_table = _pdf_substantial_table_rows(table_rows_snapshot, import_mode)
        from_cart_pdf = len(cart_rows) >= 1 and not substantial_table
        if from_cart_pdf and len(table_rows_snapshot) >= max(3, len(cart_rows) * 3):
            from_cart_pdf = False
        if from_cart_pdf:
            all_rows = cart_rows
        manual_header_layout = None if from_cart_pdf else _pdf_parse_manual_header_layout(request)
        header_layout = manual_header_layout if manual_header_layout else (
            None if from_cart_pdf else _pdf_detect_header_layout(all_rows, import_mode)
        )
        header_layout_source = 'manual' if manual_header_layout else ('auto' if header_layout else None)

        # Получаем каталог пользователя
        catalog_items = fetch_all(
            "SELECT id, name, article, brand, category, retail_price, purchase_price, wholesale_price, item_type FROM catalog_materials WHERE user_id = ?",
            (current_user.id,)
        )

        catalog_by_article = _pdf_catalog_article_map(catalog_items)
        catalog_search_index = []
        for item in catalog_items:
            norm = _pdf_normalize_name_for_index(item['name'])
            name_lower = (item['name'] or '').lower()
            if not (norm or '').strip():
                norm = name_lower
            blob = ' '.join(
                filter(
                    None,
                    [item.get('name'), item.get('brand'), item.get('article'), item.get('category')],
                )
            )
            norm_cmp = _pdf_normalize_for_compare(blob)
            catalog_search_index.append((item, norm, name_lower, norm_cmp))

        matched = []
        unmatched = []
        MIN_MATCH_SCORE = 0.22

        for row in all_rows:
            if not row or all(not cell for cell in row):
                continue

            row_text = ' '.join(str(cell) for cell in row if cell).strip()
            if len(row_text) < 3:
                continue

            if _PDF_HEADER_START.search(row_text):
                continue
            if _PDF_TOTALS_ROW.search(row_text):
                continue

            table_layout = None if from_cart_pdf else _pdf_table_layout_for_row(row, import_mode, header_layout)
            if table_layout:
                ni = table_layout.get('name_idx')
                row_name = (
                    str(row[int(ni)] or '').strip()
                    if ni is not None and int(ni) < len(row)
                    else ''
                )
            else:
                row_name = _pdf_row_product_title(row_text)
            if _pdf_is_non_product_row(row_name, row_text):
                continue

            pdf_retail_unit = _pdf_extract_pdf_retail_unit_price(row, from_cart_pdf, table_layout)
            pdf_wholesale_unit = _pdf_extract_pdf_unit_price_with_vat(row, from_cart_pdf, table_layout)
            if (
                import_mode == 'wholesale'
                and pdf_wholesale_unit is not None
                and not from_cart_pdf
                and not _pdf_is_supplier_pricelist_layout(table_layout)
            ):
                pdf_wholesale_unit = round(
                    float(pdf_wholesale_unit) * _PDF_WHOLESALE_EX_VAT_TO_WITH_VAT, 4
                )
            file_unit_price = pdf_wholesale_unit if import_mode == 'wholesale' else pdf_retail_unit

            article_match_item = _pdf_match_by_article(row_text, row, catalog_by_article)
            best_match = article_match_item
            # Совпадение по артикулу не отменяем из‑за расхождения цены PDF и каталога.
            best_score = 1.0 if best_match else 0.0
            second_score = 0.0
            from_article = bool(article_match_item)
            lex_ratio = 1.0
            word_cov = 1.0

            if table_layout:
                clean_name = str(row[table_layout['name_idx']] or '').strip()
            else:
                clean_name = _pdf_row_product_title(row_text)
            if len(clean_name) < 4:
                clean_name = row_text
                clean_name = re.sub(r'^\d+\s+', '', clean_name)
                clean_name = re.sub(r'^\d+[\.,]\d{2}\s+', '', clean_name)
                clean_name = re.sub(r'\s*В\s*наличии.*$', '', clean_name, flags=re.I)
                clean_name = clean_name.strip()

            if not best_match and clean_name:
                best_match, best_score, second_score, lex_ratio, word_cov = _pdf_fuzzy_best_match(
                    clean_name,
                    catalog_search_index,
                    (pdf_retail_unit if import_mode == 'retail' else None),
                )

            reach_match = (
                from_article
                or best_score >= MIN_MATCH_SCORE
                or (pdf_retail_unit is not None and best_score >= 0.17)
            )
            if best_match and reach_match:
                pdf_unit = _pdf_extract_unit(row, table_layout)
                raw_qty = _pdf_extract_qty(row, from_cart_pdf, table_layout)
                pdf_purchase_unit = None
                if table_layout:
                    pidx = table_layout.get('purchase_idx')
                    if pidx is not None and int(pidx) < len(row):
                        pdf_purchase_unit = _pdf_parse_money_cell(row[int(pidx)])
                if pdf_purchase_unit is None and table_layout:
                    pdf_purchase_unit = table_layout.get('purchase_from_tail')
                qty = _pdf_finalize_quantity(
                    raw_qty,
                    pdf_retail_unit,
                    pdf_purchase_unit,
                    table_layout,
                )
                qty = _pdf_sanitize_qty(
                    qty,
                    best_match.get('retail_price'),
                    best_match.get('purchase_price'),
                    file_retail=pdf_retail_unit,
                    file_wholesale=pdf_wholesale_unit,
                )
                if from_article:
                    conf_pct = 100.0
                else:
                    conf_pct = _pdf_match_confidence_display(best_score, lex_ratio, word_cov)
                ambiguous = (
                    not from_article
                    and second_score >= best_score - 0.035
                    and second_score >= 0.32
                )
                if ambiguous:
                    conf_pct = round(min(conf_pct * 0.88, 100), 1)
                # Розница из PDF не содержит закуп; не масштабируем из каталога (давало 8.352 вместо 53.38).
                file_purchase = None
                if import_mode != 'retail' and pdf_retail_unit is not None:
                    file_purchase = _pdf_purchase_for_file_retail(
                        pdf_retail_unit, best_match, qty=qty
                    )
                matched.append({
                    'row': row,
                    'row_text': row_text,
                    'clean_name': clean_name,
                    'item': {
                        'id': best_match['id'],
                        'name': best_match['name'],
                        'article': best_match['article'],
                        'brand': best_match['brand'],
                        'category': best_match['category'],
                        'retail_price': best_match['retail_price'],
                        'purchase_price': best_match['purchase_price'],
                        'wholesale_price': best_match.get('wholesale_price'),
                        'item_type': best_match['item_type'],
                    },
                    'qty': qty,
                    'unit': pdf_unit,
                    'confidence': conf_pct,
                    'ambiguous': ambiguous,
                    'file_retail_unit': round(float(pdf_retail_unit), 4)
                    if pdf_retail_unit is not None
                    else None,
                    'file_wholesale_unit': round(float(pdf_wholesale_unit), 4)
                    if pdf_wholesale_unit is not None
                    else None,
                    'file_unit_price': round(float(file_unit_price), 4)
                    if file_unit_price is not None
                    else None,
                    'file_purchase_unit': file_purchase,
                })
            else:
                if (
                    import_mode == 'retail'
                    and
                    not best_match
                    and pdf_retail_unit is not None
                    and catalog_items
                    and not any(
                        _pdf_retail_price_match(it.get('retail_price'), pdf_retail_unit)
                        for it in catalog_items
                    )
                ):
                    rej = 'price_not_in_catalog'
                elif not best_match and best_score >= 0.35 and second_score >= best_score - 0.035:
                    rej = 'ambiguous'
                elif not best_match and best_score > 0:
                    rej = 'low_score'
                else:
                    rej = 'no_match' if catalog_items else 'no_catalog'
                raw_qty_u = _pdf_extract_qty(row, from_cart_pdf, table_layout)
                qty_u = _pdf_finalize_quantity(
                    raw_qty_u,
                    pdf_retail_unit,
                    pdf_wholesale_unit if import_mode == 'wholesale' else None,
                    table_layout,
                )
                qty_u = _pdf_sanitize_qty(
                    qty_u,
                    (file_unit_price if file_unit_price is not None else 0),
                    0,
                    file_retail=pdf_retail_unit,
                    file_wholesale=pdf_wholesale_unit,
                )
                unit_u = _pdf_extract_unit(row, table_layout)
                unmatched.append({
                    'row': row,
                    'row_text': row_text,
                    'clean_name': clean_name,
                    'reject_reason': rej,
                    'qty': qty_u,
                    'unit': unit_u,
                    'file_unit_price': round(float(file_unit_price), 4)
                    if file_unit_price is not None
                    else None,
                })

        # Для UI сопоставления колонок: реальная ширина таблицы часто больше, чем строка заголовка.
        max_cols = 0
        for r in all_rows:
            if not r:
                continue
            max_cols = max(max_cols, len(r))
        if header_layout and isinstance(header_layout, dict) and max_cols > 0:
            cells = list(header_layout.get('header_cells') or [])
            if len(cells) < max_cols:
                for i in range(len(cells), max_cols):
                    cells.append(f'Колонка {i + 1}')
                header_layout = {**header_layout, 'header_cells': cells}
            header_layout = {**header_layout, 'max_cols': max_cols}

        return jsonify({
            "import_mode": import_mode,
            "header_layout": header_layout,
            "header_layout_source": header_layout_source,
            "matched": matched,
            "unmatched": unmatched,
            "total_rows": len(all_rows),
            "matched_count": len(matched),
            "unmatched_count": len(unmatched),
            "catalog_count": len(catalog_items),
            "catalog_empty": len(catalog_items) == 0,
        })

    except Exception as e:
        import logging, traceback
        tb = traceback.format_exc()
        logging.error(f"parse_pdf error: {e}\n{tb}")
        return jsonify({"error": str(e)}), 500


@estimate_bp.route('/api/pdf-to-sheet', methods=['POST'])
@login_required
def api_pdf_to_sheet():
    """Преобразовать PDF в полную Excel-таблицу (без урезания строк/колонок)."""
    try:
        import pdfplumber

        if 'file' not in request.files:
            return jsonify({"error": "Нет файла"}), 400

        file = request.files['file']
        if not file or not (file.filename or '').lower().endswith('.pdf'):
            return jsonify({"error": "Нужен PDF файл"}), 400

        mode_raw = (request.form.get('mode') or request.args.get('mode') or 'retail').strip().lower()
        import_mode = 'wholesale' if mode_raw in ('wholesale', 'opt', 'purchase', 'cost') else 'retail'

        pdf_file = pdfplumber.open(io.BytesIO(file.read()))
        full_text_chunks = []
        all_rows = []
        for page in pdf_file.pages:
            text = page.extract_text()
            if text:
                full_text_chunks.append(text)
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if row and any(cell for cell in row if cell):
                            all_rows.append([str(c).strip() if c else '' for c in row])
            elif text:
                for line in text.split('\n'):
                    line = line.strip()
                    if not line:
                        continue
                    parts = re.split(r'\t|\s{3,}', line)
                    parsed = [p.strip() for p in parts if p.strip()]
                    if parsed:
                        all_rows.append(parsed)
        pdf_file.close()

        all_rows = _pdf_filter_table_rows(all_rows)

        full_text = '\n'.join(full_text_chunks)
        cart_rows = _pdf_parse_cart_ready_format(full_text)
        from_cart_pdf = len(cart_rows) >= 1
        if from_cart_pdf:
            all_rows = cart_rows

        if not all_rows:
            return jsonify({"error": "Не удалось извлечь таблицу из PDF"}), 400

        manual_header_layout = None if from_cart_pdf else _pdf_parse_manual_header_layout(request)
        header_layout = manual_header_layout if manual_header_layout else (
            None if from_cart_pdf else _pdf_detect_header_layout(all_rows, import_mode)
        )

        max_cols = 0
        for row in all_rows:
            if row:
                max_cols = max(max_cols, len(row))
        if max_cols <= 0:
            return jsonify({"error": "В PDF нет табличных данных"}), 400

        header_cells = []
        src_cells = list((header_layout or {}).get('header_cells') or [])
        for i in range(max_cols):
            cell = src_cells[i] if i < len(src_cells) else ''
            header_cells.append((str(cell).strip() if cell else '') or f'Колонка {i + 1}')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Опт' if import_mode == 'wholesale' else 'Розница'

        ws.append(header_cells)
        for row in all_rows:
            padded = list(row) + [''] * max(0, max_cols - len(row))
            ws.append(padded[:max_cols])

        # Лёгкое авто-растяжение колонок для удобного просмотра.
        for col_idx in range(1, max_cols + 1):
            max_len = len(str(header_cells[col_idx - 1] or ''))
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 80)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        suffix = 'opt' if import_mode == 'wholesale' else 'retail'
        filename = f'pdf_{suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            out,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logging.error(f"pdf_to_sheet error: {e}\n{tb}")
        return jsonify({"error": str(e)}), 500


def _pdf_extract_rows_for_sheet(uploaded_file):
    import pdfplumber
    pdf_file = pdfplumber.open(io.BytesIO(uploaded_file.read()))
    full_text_chunks = []
    all_rows = []
    for page in pdf_file.pages:
        text = page.extract_text()
        if text:
            full_text_chunks.append(text)
        tables = page.extract_tables()
        if tables:
            for table in tables:
                for row in table:
                    if row and any(cell for cell in row if cell):
                        all_rows.append([str(c).strip() if c else '' for c in row])
        elif text:
            for line in text.split('\n'):
                line = line.strip()
                if not line:
                    continue
                parts = re.split(r'\t|\s{3,}', line)
                parsed = [p.strip() for p in parts if p.strip()]
                if parsed:
                    all_rows.append(parsed)
    pdf_file.close()
    all_rows = _pdf_filter_table_rows(all_rows)
    full_text = '\n'.join(full_text_chunks)
    cart_rows = _pdf_parse_cart_ready_format(full_text)
    if len(cart_rows) >= 1:
        return cart_rows
    return all_rows


def _to_float_ru(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace('\u00a0', ' ').replace(' ', '').replace(',', '.')
    s = re.sub(r'[^0-9.\-]', '', s)
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _pdf_supplier_line_number(row, table_layout=None):
    """Номер позиции в счёте/прайсе (колонка №)."""
    try:
        if not row:
            return None
        if table_layout and table_layout.get('layout_source') == 'supplier_pricelist':
            num_idx = int(table_layout.get('name_idx', 2)) - 1
            if 0 <= num_idx < len(row):
                s = str(row[num_idx] or '').strip()
                if re.fullmatch(r'\d{1,4}', s):
                    return int(s)
        for idx in (1, 0):
            if idx >= len(row):
                continue
            s = str(row[idx] or '').strip()
            if re.fullmatch(r'\d{1,4}', s):
                return int(s)
    except (TypeError, ValueError):
        pass
    return None


def _pdf_wholesale_unit_from_row(wrow, wholesale_header_layout):
    """Цена закупа за ед. из строки оптового PDF."""
    w_layout = _pdf_table_layout_for_row(wrow, 'wholesale', wholesale_header_layout)
    if w_layout is None:
        w_layout = _pdf_detect_supplier_pricelist_row_layout(wrow, 'wholesale')
    wholesale_price = _pdf_extract_wholesale_unit_supplier(wrow, w_layout)
    if wholesale_price is None and w_layout and w_layout.get('retail_idx') is not None:
        try:
            ri = int(w_layout['retail_idx'])
            if 0 <= ri < len(wrow):
                wholesale_price = _to_float_ru(wrow[ri])
        except (TypeError, ValueError):
            pass
    if wholesale_price is None and len(wrow) > 5:
        wholesale_price = _to_float_ru(wrow[5])
    if (
        wholesale_price is not None
        and not _pdf_is_supplier_pricelist_layout(w_layout)
    ):
        wholesale_price = round(
            float(wholesale_price) * _PDF_WHOLESALE_EX_VAT_TO_WITH_VAT, 4
        )
    return wholesale_price


def _pdf_build_wholesale_lookup(wholesale_rows, wholesale_header_layout):
    """Индекс оптовых строк: по № позиции и по нормализованному наименованию."""
    by_line = {}
    by_name = {}
    for row in wholesale_rows or []:
        row_text = ' '.join(str(c or '') for c in row).strip()
        if len(row_text) < 3:
            continue
        if _PDF_HEADER_START.search(row_text) or _PDF_TOTALS_ROW.search(row_text):
            continue
        w_layout = _pdf_table_layout_for_row(row, 'wholesale', wholesale_header_layout)
        if w_layout is None:
            w_layout = _pdf_detect_supplier_pricelist_row_layout(row, 'wholesale')
        if not _pdf_is_supplier_pricelist_layout(w_layout):
            continue
        price = _pdf_wholesale_unit_from_row(row, wholesale_header_layout)
        if price is None or price <= 0:
            continue
        line_no = _pdf_supplier_line_number(row, w_layout)
        if line_no is not None:
            by_line[line_no] = float(price)
        try:
            ni = int(w_layout['name_idx'])
            name = str(row[ni] or '').strip()
        except (TypeError, ValueError, KeyError, IndexError):
            name = ''
        if len(name) >= 3:
            nk = _pdf_normalize_name_for_index(name)
            if nk:
                by_name.setdefault(nk, []).append(float(price))
    return by_line, by_name


def _pdf_match_wholesale_for_retail(retail_row, retail_layout, wholesale_by_line, wholesale_by_name):
    """Сопоставить строку РРЦ со строкой счёта по № или наименованию (не по индексу массива)."""
    line_no = _pdf_supplier_line_number(retail_row, retail_layout)
    if line_no is not None and line_no in wholesale_by_line:
        return wholesale_by_line[line_no]
    try:
        ni = int(retail_layout['name_idx'])
        rname = str(retail_row[ni] or '').strip()
    except (TypeError, ValueError, KeyError, IndexError):
        rname = ''
    nk = _pdf_normalize_name_for_index(rname)
    if nk and nk in wholesale_by_name:
        return wholesale_by_name[nk][0]
    if len(nk) >= 10:
        for key, prices in wholesale_by_name.items():
            if nk in key or key in nk:
                return prices[0]
    return None


@estimate_bp.route('/api/pdf-merge-retail-wholesale-sheet', methods=['POST'])
@login_required
def api_pdf_merge_retail_wholesale_sheet():
    """
    Объединение двух PDF в одну таблицу:
    - база: розничная таблица;
    - 7-я колонка: заголовок "Опт";
    - ниже: значения из 6-й колонки оптового PDF, увеличенные на 20%.
    """
    try:
        retail_file = request.files.get('retail_file')
        wholesale_file = request.files.get('wholesale_file')
        if not retail_file or not wholesale_file:
            return jsonify({"error": "Нужны два файла: retail_file и wholesale_file"}), 400
        if not (retail_file.filename or '').lower().endswith('.pdf'):
            return jsonify({"error": "Файл розницы должен быть PDF"}), 400
        if not (wholesale_file.filename or '').lower().endswith('.pdf'):
            return jsonify({"error": "Файл опта должен быть PDF"}), 400

        retail_rows = _pdf_extract_rows_for_sheet(retail_file)
        wholesale_rows = _pdf_extract_rows_for_sheet(wholesale_file)
        if not retail_rows:
            return jsonify({"error": "Не удалось извлечь таблицу из PDF розницы"}), 400

        retail_mode = 'retail'
        retail_header_layout = _pdf_detect_header_layout(retail_rows, retail_mode)
        wholesale_header_layout = (
            _pdf_detect_header_layout(wholesale_rows, 'wholesale') if wholesale_rows else None
        )
        wholesale_by_line, wholesale_by_name = _pdf_build_wholesale_lookup(
            wholesale_rows, wholesale_header_layout
        )

        retail_max_cols = 0
        for row in retail_rows:
            if row:
                retail_max_cols = max(retail_max_cols, len(row))
        if retail_max_cols < 7:
            retail_max_cols = 7

        header_cells = []
        src_cells = list((retail_header_layout or {}).get('header_cells') or [])
        for i in range(retail_max_cols):
            cell = src_cells[i] if i < len(src_cells) else ''
            header_cells.append((str(cell).strip() if cell else '') or f'Колонка {i + 1}')
        header_cells[6] = 'Опт'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Розница+Опт'
        ws.append(header_cells)

        for retail_row in retail_rows:
            row_out = list(retail_row) + [''] * max(0, retail_max_cols - len(retail_row))
            row_out = row_out[:retail_max_cols]

            retail_layout = _pdf_table_layout_for_row(retail_row, 'retail', retail_header_layout)
            if retail_layout is None:
                retail_layout = _pdf_detect_supplier_pricelist_row_layout(retail_row, 'retail')
            wholesale_num = None
            if _pdf_is_supplier_pricelist_layout(retail_layout):
                wholesale_num = _pdf_match_wholesale_for_retail(
                    retail_row, retail_layout, wholesale_by_line, wholesale_by_name
                )
            row_out[6] = wholesale_num if wholesale_num is not None else ''

            ws.append(row_out)

        for col_idx in range(1, retail_max_cols + 1):
            max_len = len(str(header_cells[col_idx - 1] or ''))
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 80)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        filename = f'pdf_merged_retail_opt_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            out,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logging.error(f"pdf_merge_retail_wholesale_sheet error: {e}\n{tb}")
        return jsonify({"error": str(e)}), 500


@estimate_bp.route('/api/pdf-import-dual-materials', methods=['POST'])
@login_required
def api_pdf_import_dual_materials():
    """
    Импорт материалов из двух PDF (розница + опт):
    - формирует единый список материалов,
    - сверяет с каталогом,
    - добавляет отсутствующие,
    - добавляет материалы в смету (если передан estimate_id).
    """
    try:
        retail_file = request.files.get('retail_file')
        wholesale_file = request.files.get('wholesale_file')
        if not retail_file or not wholesale_file:
            return jsonify({"error": "Нужны два файла: retail_file и wholesale_file"}), 400
        if not (retail_file.filename or '').lower().endswith('.pdf'):
            return jsonify({"error": "Файл розницы должен быть PDF"}), 400
        if not (wholesale_file.filename or '').lower().endswith('.pdf'):
            return jsonify({"error": "Файл опта должен быть PDF"}), 400

        estimate_id_raw = (request.form.get('estimate_id') or '').strip()
        estimate_id = int(estimate_id_raw) if estimate_id_raw.isdigit() else None
        if estimate_id is not None:
            est = fetch_one("SELECT id FROM estimates WHERE id = ? AND user_id = ?", (estimate_id, current_user.id))
            if not est:
                return jsonify({"error": "Смета не найдена"}), 404

        retail_rows = _pdf_extract_rows_for_sheet(retail_file)
        wholesale_rows = _pdf_extract_rows_for_sheet(wholesale_file)
        if not retail_rows:
            return jsonify({"error": "Не удалось извлечь таблицу из PDF розницы"}), 400

        retail_header_layout = _pdf_detect_header_layout(retail_rows, 'retail')
        wholesale_header_layout = _pdf_detect_header_layout(wholesale_rows, 'wholesale') if wholesale_rows else None
        wholesale_by_line, wholesale_by_name = _pdf_build_wholesale_lookup(
            wholesale_rows, wholesale_header_layout
        )

        catalog_items = fetch_all(
            "SELECT id, name, article, brand, category, item_type, retail_price, purchase_price, wholesale_price FROM catalog_materials WHERE user_id = ?",
            (current_user.id,),
        )
        catalog_by_norm_name = {}
        for it in catalog_items:
            key = _pdf_normalize_name_for_index(it.get('name') or '')
            if key and key not in catalog_by_norm_name:
                catalog_by_norm_name[key] = it

        materials = []
        for row in retail_rows:
            if not row or all(not str(c or '').strip() for c in row):
                continue
            row_text = ' '.join(str(cell or '') for cell in row).strip()
            if len(row_text) < 3:
                continue
            if _PDF_HEADER_START.search(row_text) or _PDF_TOTALS_ROW.search(row_text):
                continue

            retail_layout = _pdf_table_layout_for_row(row, 'retail', retail_header_layout)
            if retail_layout is None:
                retail_layout = _pdf_detect_supplier_pricelist_row_layout(row, 'retail')
            if not _pdf_is_supplier_pricelist_layout(retail_layout):
                continue
            name = str(row[int(retail_layout['name_idx'])] or '').strip()
            if not name or len(name) < 3:
                continue

            qty = None
            if retail_layout and retail_layout.get('qty_idx') is not None:
                q_idx = int(retail_layout.get('qty_idx'))
                if 0 <= q_idx < len(row):
                    qty = _to_float_ru(row[q_idx])
            if qty is None and len(row) > 2:
                # Страховка для частого формата счетов: количество в 3-й колонке.
                qty = _to_float_ru(row[2])
            if qty is None:
                qty = _pdf_sanitize_qty(_pdf_extract_qty(row, False, retail_layout), 0, 0)
            qty = float(qty or 0)
            if qty <= 0:
                qty = 1.0

            unit = _pdf_extract_unit(row, retail_layout)
            if retail_layout and retail_layout.get('unit_idx') is not None:
                u_idx = int(retail_layout.get('unit_idx'))
                if 0 <= u_idx < len(row):
                    unit = _pdf_unit_from_table_cell(row[u_idx])
            unit = _normalize_estimate_unit(unit, 'material')

            retail_price = None
            if retail_layout and retail_layout.get('retail_idx') is not None:
                r_idx = int(retail_layout.get('retail_idx'))
                if 0 <= r_idx < len(row):
                    retail_price = _to_float_ru(row[r_idx])
            if retail_price is None and len(row) > 4:
                retail_price = _to_float_ru(row[4])
            if retail_price is None:
                retail_price = _pdf_extract_pdf_retail_unit_price(row, False, retail_layout)

            wholesale_price = _pdf_match_wholesale_for_retail(
                row, retail_layout, wholesale_by_line, wholesale_by_name
            )

            retail_price = float(retail_price or 0)
            wholesale_price = float(wholesale_price or 0)
            if retail_price <= 0 and wholesale_price <= 0:
                continue
            if retail_price <= 0 and wholesale_price > 0:
                retail_price = wholesale_price

            materials.append({
                'name': name,
                'unit': unit,
                'qty': float(qty or 1),
                'retail_price': round(retail_price, 4),
                'wholesale_price': round(wholesale_price, 4),
            })

        if not materials:
            retail_n = sum(
                1
                for row in retail_rows
                if row
                and _pdf_is_supplier_pricelist_layout(
                    _pdf_detect_supplier_pricelist_row_layout(row, 'retail')
                    or _pdf_table_layout_for_row(row, 'retail', retail_header_layout)
                )
            )
            wholesale_n = sum(
                1
                for row in (wholesale_rows or [])
                if row
                and _pdf_is_supplier_pricelist_layout(
                    _pdf_detect_supplier_pricelist_row_layout(row, 'wholesale')
                    or _pdf_table_layout_for_row(row, 'wholesale', wholesale_header_layout)
                )
            )
            hint = (
                f"Розница: распознано строк {retail_n}, опт: {wholesale_n}. "
                "Проверьте: в «PDF розница» — счёт с ценами для клиента, в «PDF опт» — прайс РРЦ. "
                "Файлы не перепутаны?"
            )
            return jsonify({"error": "Не удалось собрать материалы из двух PDF. " + hint}), 400

        added_to_catalog = 0
        updated_in_catalog = 0
        added_to_estimate = 0
        updated_in_estimate = 0

        estimate_items = []
        if estimate_id is not None:
            estimate_items = fetch_all(
                """SELECT id, section, name, unit, quantity, price, purchase_price, wholesale_price
                   FROM estimate_items WHERE estimate_id = ?""",
                (estimate_id,),
            )

        for m in materials:
            norm_key = _pdf_normalize_name_for_index(m['name'])
            existing = catalog_by_norm_name.get(norm_key) if norm_key else None
            if existing:
                execute(
                    """UPDATE catalog_materials
                       SET retail_price = ?, purchase_price = ?, wholesale_price = ?, unit = ?, use_count = COALESCE(use_count,0)+1
                       WHERE id = ? AND user_id = ?""",
                    (m['retail_price'], m['wholesale_price'], m['wholesale_price'], m['unit'], existing['id'], current_user.id),
                )
                catalog_id = existing['id']
                updated_in_catalog += 1
            else:
                catalog_id = execute(
                    """INSERT INTO catalog_materials
                       (user_id, name, unit, category, article, brand, item_type, purchase_price, retail_price, wholesale_price, min_wholesale_qty, description, use_count)
                       VALUES (?, ?, ?, '', '', '', 'material', ?, ?, ?, 10, '', 1)""",
                    (current_user.id, m['name'], m['unit'], m['wholesale_price'], m['retail_price'], m['wholesale_price']),
                    return_id=True,
                )
                added_to_catalog += 1
                if norm_key:
                    catalog_by_norm_name[norm_key] = {'id': catalog_id, 'name': m['name']}

            if estimate_id is not None:
                total = m['retail_price'] * m['qty']
                profit = (m['retail_price'] - m['wholesale_price']) * m['qty']
                matches = _estimate_match_material_items(
                    estimate_items, m['name'], m['retail_price']
                )
                if matches:
                    for it in matches:
                        pu = _normalize_unit_purchase_price(
                            m['wholesale_price'], m['retail_price'], qty=m['qty']
                        )
                        execute(
                            """UPDATE estimate_items
                               SET unit=?, quantity=?, price=?, purchase_price=?, wholesale_price=?,
                                   total=?, material_profit=?
                               WHERE id=? AND estimate_id=?""",
                            (
                                m['unit'],
                                m['qty'],
                                m['retail_price'],
                                pu,
                                m['wholesale_price'],
                                total,
                                profit,
                                it['id'],
                                estimate_id,
                            ),
                        )
                        it['price'] = m['retail_price']
                        it['purchase_price'] = pu
                        it['quantity'] = m['qty']
                    updated_in_estimate += len(matches)
                else:
                    execute(
                        """INSERT INTO estimate_items
                           (estimate_id, section, name, unit, quantity, price_type, price, purchase_price, wholesale_price, total, material_profit, sort_order)
                           VALUES (?, 'material', ?, ?, ?, 'retail', ?, ?, ?, ?, ?, (SELECT COALESCE(MAX(sort_order),0)+1 FROM estimate_items WHERE estimate_id=?))""",
                        (
                            estimate_id,
                            m['name'],
                            m['unit'],
                            m['qty'],
                            m['retail_price'],
                            m['wholesale_price'],
                            m['wholesale_price'],
                            total,
                            profit,
                            estimate_id,
                        ),
                    )
                    added_to_estimate += 1

        return jsonify({
            "ok": True,
            "materials_count": len(materials),
            "added_to_catalog": added_to_catalog,
            "updated_in_catalog": updated_in_catalog,
            "added_to_estimate": added_to_estimate,
            "updated_in_estimate": updated_in_estimate,
            "estimate_id": estimate_id,
        })
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logging.error(f"pdf_import_dual_materials error: {e}\n{tb}")
        return jsonify({"error": str(e)}), 500

def _opt_config_path():
    import os
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), '.opt_config.json')


@estimate_bp.route('/api/price-sync/checkpoint', methods=['GET'])
@login_required
def price_sync_checkpoint():
    """Чекпоинт: что сохранено для opt-akvabreg (логин и длина/маска пароля, без раскрытия)."""
    import json, os
    path = _opt_config_path()
    if not os.path.exists(path):
        return jsonify({
            "configured": False,
            "login": "",
            "password_saved": False,
            "password_length": 0,
            "password_mask": "",
        })
    try:
        with open(path, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
    except Exception as e:
        return jsonify({"error": str(e), "configured": False}), 500
    login_v = (cfg.get('login') or '').strip()
    pwd = cfg.get('password') or ''
    if not isinstance(pwd, str):
        pwd = str(pwd)
    n = len(pwd)
    mask = ('\u2022' * min(n, 40) + ('\u2026' if n > 40 else '')) if n else ''
    return jsonify({
        "configured": bool(login_v or pwd),
        "login": login_v,
        "password_saved": bool(pwd),
        "password_length": n,
        "password_mask": mask,
    })


@estimate_bp.route('/api/price-sync/reveal-password', methods=['POST'])
@login_required
@_require_csrf
def price_sync_reveal_password():
    """Показать сохранённый пароль в интерфейсе (только для текущего пользователя приложения)."""
    import json, os
    path = _opt_config_path()
    if not os.path.exists(path):
        return jsonify({"error": "Конфиг не найден", "password": ""}), 404
    try:
        with open(path, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"password": cfg.get('password') or ""})


@estimate_bp.route('/api/price-sync/config', methods=['POST'])
@login_required
@_require_csrf
def price_sync_config():
    """Сохранить логин/пароль от opt-akvabreg.by (пустой пароль не затирает уже сохранённый)."""
    data, err = _require_json()
    if err: return err
    try:
        import json, os
        config_path = _opt_config_path()
        existing = {}
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    existing = json.load(f)
            except Exception:
                existing = {}
        login = (data.get('login') if 'login' in data else None)
        if login is None:
            login = existing.get('login', '')
        login = (login or '').strip()
        password = data.get('password', '')
        if password is None:
            password = ''
        if isinstance(password, str) and password.strip() == '':
            password = existing.get('password', '')
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump({'login': login, 'password': password}, f, ensure_ascii=False)
        return jsonify({"status": "ok"})
    except Exception as e:
        logger.error(f"price_sync_config error: {e}")
        return jsonify({"error": str(e)}), 500

@estimate_bp.route('/api/price-sync/run', methods=['POST'])
@login_required
@_require_csrf
def price_sync_run():
    """Запустить синхронизацию цен"""
    try:
        from price_sync import login, compare_prices, download_price_list, parse_price_list_dataframe
        import pandas as pd, os

        session, msg = login()
        if not session:
            return jsonify({"error": msg, "status": "auth_failed"}), 401

        filepath, filename, msg = download_price_list(session)
        if not filepath:
            return jsonify({"error": msg, "status": "download_failed"}), 500

        # Загрузка локальных цен
        local_items = fetch_all(
            "SELECT id, name, article, retail_price, purchase_price FROM catalog_materials WHERE user_id = ?",
            (current_user.id,)
        )

        df = pd.read_excel(filepath)
        new_items = parse_price_list_dataframe(df)

        results = compare_prices(local_items, new_items)
        results['filepath'] = filepath
        results['filename'] = filename

        return jsonify(results)

    except Exception as e:
        import logging, traceback
        tb = traceback.format_exc()
        logging.error(f"price_sync_run error: {e}\n{tb}")
        return jsonify({"error": str(e)}), 500

@estimate_bp.route('/api/price-sync/apply', methods=['POST'])
@login_required
@_require_csrf
def price_sync_apply():
    """Применить новые цены из результата синхронизации"""
    data, err = _require_json()
    if err: return err
    try:
        updates = data.get('updates', [])
        updated = 0
        for item in updates:
            article = item.get('article', '').strip().upper()
            new_price = float(item.get('new_price', 0))
            if article and new_price > 0:
                execute(
                    "UPDATE catalog_materials SET retail_price = ? WHERE user_id = ? AND UPPER(article) = ?",
                    (new_price, current_user.id, article)
                )
                updated += 1
        return jsonify({"status": "ok", "updated": updated})
    except Exception as e:
        logger.error(f"price_sync_apply error: {e}")
        return jsonify({"error": str(e)}), 500

@estimate_bp.route('/api/catalog/categories/<int:cat_id>', methods=['DELETE'])
@login_required
@_require_csrf
def api_delete_category(cat_id):
    execute("DELETE FROM categories WHERE id = ? AND user_id = ?", (cat_id, current_user.id))
    return jsonify({"ok": True})

# ============================
# API: СМЕТЫ
# ============================

def _next_estimate_number(user_id):
    """Следующий номер С-NNNN по максимуму среди существующих (данные не трогаем)."""
    rows = fetch_all("SELECT number FROM estimates WHERE user_id = ?", (user_id,))
    max_n = 0
    for r in rows:
        num = (r.get('number') or '').strip()
        m = re.match(r'^С-(\d+)$', num, re.IGNORECASE)
        if m:
            max_n = max(max_n, int(m.group(1)))
    return f"С-{max_n + 1:04d}"


def _resolve_object_id_for_user(raw, user_id):
    """
    object_id из тела запроса: 0 / отсутствует — без привязки к объекту.
    Иначе только если объект существует и принадлежит user_id.
    """
    if raw is None or raw == '':
        return 0, None
    try:
        oid = int(raw)
    except (TypeError, ValueError):
        return None, (jsonify({"error": "Некорректный object_id"}), 400)
    if oid <= 0:
        return 0, None
    if not fetch_one("SELECT 1 AS x FROM objects WHERE id = ? AND user_id = ?", (oid, user_id)):
        return None, (jsonify({"error": "Объект не найден или не принадлежит вам"}), 400)
    return oid, None


@estimate_bp.route('/api/estimates', methods=['GET'])
@login_required
def api_get_estimates():
    estimates = fetch_all("SELECT * FROM estimates WHERE user_id = ? ORDER BY date DESC", (current_user.id,))
    return jsonify(estimates)

@estimate_bp.route('/api/estimates', methods=['POST'])
@login_required
@_require_csrf
def api_create_estimate():
    data, err = _require_json()
    if err: return err
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    number = _next_estimate_number(current_user.id)

    object_id, bad = _resolve_object_id_for_user(data.get('object_id'), current_user.id)
    if bad:
        return bad

    eid = execute("""INSERT INTO estimates
        (user_id, number, date, object_id, object_name, client, status, vat_percent, markup_percent, discount_percent, notes, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (current_user.id, number, data.get('date', datetime.now().strftime('%Y-%m-%d')),
         object_id, data.get('object_name', ''), data.get('client', ''),
         data.get('status', 'Черновик'), _safe_float(data.get('vat_percent')), _safe_float(data.get('markup_percent')),
         _safe_float(data.get('discount_percent')), data.get('notes', ''), now, now), return_id=True)
    return jsonify({"id": eid, "number": number}), 201

@estimate_bp.route('/api/estimates/<int:est_id>', methods=['GET'])
@login_required
def api_get_estimate(est_id):
    est = fetch_one("SELECT * FROM estimates WHERE id = ? AND user_id = ?", (est_id, current_user.id))
    if not est: return jsonify({"error": "Not found"}), 404
    items = fetch_all("SELECT * FROM estimate_items WHERE estimate_id = ?", (est_id,))
    _backfill_estimate_wholesale_from_catalog(current_user.id, est_id, items)
    for it in items:
        sec = it.get('section') or 'material'
        nu = _normalize_estimate_unit(it.get('unit'), sec)
        if nu != (it.get('unit') or ''):
            execute(
                'UPDATE estimate_items SET unit = ? WHERE id = ? AND estimate_id = ?',
                (nu, it['id'], est_id),
            )
            it['unit'] = nu
        if sec == 'material':
            q = _safe_float(it.get('quantity'), default=1.0)
            p = _safe_float(it.get('price'))
            if _repair_estimate_item_quantity(it):
                execute(
                    """UPDATE estimate_items SET quantity=?, purchase_price=?, total=?, material_profit=?
                       WHERE id=? AND estimate_id=?""",
                    (
                        it['quantity'],
                        it['purchase_price'],
                        it['total'],
                        it['material_profit'],
                        it['id'],
                        est_id,
                    ),
                )
                q = _safe_float(it.get('quantity'), default=1.0)
                p = _safe_float(it.get('price'))
            pu = _normalize_unit_purchase_price(_safe_float(it.get('purchase_price')), p, qty=q)
            if pu != _safe_float(it.get('purchase_price')):
                it['purchase_price'] = pu
                it['material_profit'] = (p - pu) * q
                it['total'] = p * q
                execute(
                    """UPDATE estimate_items SET purchase_price=?, total=?, material_profit=?
                       WHERE id=? AND estimate_id=?""",
                    (pu, it['total'], it['material_profit'], it['id'], est_id),
                )
    return jsonify({**est, 'items': items})

def _dispatcher_notify_estimate_status(user_id, object_id, status, estimate_number='', customer_phone=''):
    """Webhook в диспетчер: смета «Отправлена» / «Утверждена» → стадия заявки."""
    base = (os.environ.get('DISPATCHER_API_BASE_URL') or os.environ.get('DISPATCHER_API_URL') or '').strip().rstrip('/')
    key = (os.environ.get('INTEGRATION_API_KEY') or '').strip()
    client_id = (os.environ.get('DISPATCHER_BUSINESS_CLIENT_ID') or '').strip()
    if not base or not key or not client_id:
        return
    status = (status or '').strip()
    if status not in ('Отправлена', 'Утверждена'):
        return
    task_id = ''
    phone = (customer_phone or '').strip()
    if object_id:
        obj = fetch_one(
            'SELECT integration_source, client FROM objects WHERE id = ? AND user_id = ?',
            (object_id, user_id),
        )
        if obj:
            src = str(obj.get('integration_source') or '').strip()
            if src.startswith('taskmgr:'):
                task_id = src.split(':', 1)[1].strip()
            if not phone:
                phone = str(obj.get('client') or '').strip()
    payload = {
        'clientId': client_id,
        'estimateStatus': status,
        'taskId': task_id or None,
        'customerPhone': phone or None,
        'estimateNumber': (estimate_number or '').strip() or None,
    }
    payload = {k: v for k, v in payload.items() if v is not None}
    url = f"{base}/v1/integration/object-accounting/estimate-status"
    body = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    req = urllib.request.Request(
        url=url,
        data=body,
        headers={
            'Authorization': f'Bearer {key}',
            'Content-Type': 'application/json',
            'Accept': 'application/json',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=12):
            return
    except Exception:
        logger.exception('integration: dispatcher estimate-status POST failed')

@estimate_bp.route('/api/estimates/<int:est_id>', methods=['PUT'])
@login_required
@_require_csrf
def api_update_estimate(est_id):
    data, err = _require_json()
    if err: return err
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    prev = fetch_one(
        """SELECT e.object_id, e.status AS prev_status, e.number,
                  o.integration_source, o.client
           FROM estimates e
           LEFT JOIN objects o ON o.id = e.object_id AND o.user_id = e.user_id
           WHERE e.id = ? AND e.user_id = ?""",
        (est_id, current_user.id),
    )
    if not prev:
        return jsonify({"error": "Not found"}), 404
    object_id = prev.get('object_id') or 0
    prev_status = (prev.get('prev_status') or '').strip()
    if 'object_id' in data:
        object_id, bad = _resolve_object_id_for_user(data.get('object_id'), current_user.id)
        if bad:
            return bad
    new_status = (data.get('status') or prev_status or 'Черновик').strip()
    n = execute_rowcount("""UPDATE estimates SET date=?, object_name=?, client=?, status=?,
           vat_percent=?, markup_percent=?, discount_percent=?, notes=?, object_id=?, updated_at=?
           WHERE id=? AND user_id=?""",
        (data.get('date'), data.get('object_name'), data.get('client'), new_status,
         _safe_float(data.get('vat_percent')), _safe_float(data.get('markup_percent')),
         _safe_float(data.get('discount_percent')), data.get('notes', ''), object_id, now, est_id, current_user.id))
    if n == 0:
        return jsonify({"error": "Not found"}), 404
    if new_status != prev_status and new_status in ('Отправлена', 'Утверждена'):
        _dispatcher_notify_estimate_status(
            current_user.id,
            object_id,
            new_status,
            estimate_number=prev.get('number') or '',
            customer_phone=(data.get('client') or prev.get('client') or ''),
        )
    return jsonify({"ok": True})

@estimate_bp.route('/api/estimates/<int:est_id>', methods=['DELETE'])
@login_required
@_require_csrf
def api_delete_estimate(est_id):
    # Только позиции сметы текущего пользователя (иначе IDOR: чужой est_id удалил бы чужие строки)
    execute(
        "DELETE FROM estimate_items WHERE estimate_id IN (SELECT id FROM estimates WHERE id = ? AND user_id = ?)",
        (est_id, current_user.id),
    )
    execute("DELETE FROM estimates WHERE id = ? AND user_id = ?", (est_id, current_user.id))
    return jsonify({"ok": True})

# ============================
# API: СТРОКИ СМЕТЫ
# ============================

@estimate_bp.route('/api/estimates/<int:est_id>/items', methods=['POST'])
@login_required
@_require_csrf
def api_add_item(est_id):
    data, err = _require_json()
    if err: return err
    est = fetch_one("SELECT id FROM estimates WHERE id = ? AND user_id = ?", (est_id, current_user.id))
    if not est:
        logger.warning(f"api_add_item: estimate {est_id} not found for user {current_user.id}")
        return jsonify({"error": "Forbidden"}), 403

    price = _safe_float(data.get('price'))
    purchase = _safe_float(data.get('purchase_price'))
    wholesale = _safe_float(data.get('wholesale_price')) if data.get('section') == 'material' else 0
    qty = _safe_float(data.get('quantity'), default=1.0)
    if data.get('section') == 'material':
        purchase = _normalize_unit_purchase_price(purchase, price, qty=qty)
    total = price * qty
    profit = (price - purchase) * qty if data.get('section') == 'material' else 0

    logger.info(f"api_add_item: estimate_id={est_id}, section={data.get('section')}, name={data.get('name')}, qty={qty}, price={price}")

    unit_norm = _normalize_estimate_unit(data.get('unit'), data.get('section'))
    _add_to_catalog(data.get('section'), data.get('name'), unit_norm, price, purchase)

    item_id = execute("""INSERT INTO estimate_items
        (estimate_id, section, name, unit, quantity, price_type, price, purchase_price, wholesale_price, total, material_profit, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, (SELECT COALESCE(MAX(sort_order),0)+1 FROM estimate_items WHERE estimate_id=?))""",
        (est_id, data.get('section'), data.get('name'), unit_norm, qty, 'retail', price, purchase, wholesale, total, profit, est_id), return_id=True)
    return jsonify({"id": item_id, "total": total, "material_profit": profit}), 201

@estimate_bp.route('/api/items/<int:item_id>', methods=['PUT'])
@login_required
@_require_csrf
def api_update_item(item_id):
    data, err = _require_json()
    if err: return err
    price = _safe_float(data.get('price'))
    purchase = _safe_float(data.get('purchase_price'))
    wholesale = _safe_float(data.get('wholesale_price')) if data.get('section', 'material') == 'material' else 0
    qty = _safe_float(data.get('quantity'), default=1.0)
    if data.get('section', 'material') == 'material':
        purchase = _normalize_unit_purchase_price(purchase, price, qty=qty)
    total = price * qty
    profit = (price - purchase) * qty if data.get('section', 'material') == 'material' else 0

    unit_norm = _normalize_estimate_unit(data.get('unit'), data.get('section', 'material'))
    n = execute_rowcount("""UPDATE estimate_items SET name=?, unit=?, quantity=?, price=?, purchase_price=?, wholesale_price=?, total=?, material_profit=?
               WHERE id=? AND estimate_id IN (SELECT id FROM estimates WHERE user_id=?)""",
            (data.get('name'), unit_norm, qty, price, purchase, wholesale, total, profit, item_id, current_user.id))
    if n == 0:
        return jsonify({"error": "Not found"}), 404
    return jsonify({"ok": True, "total": total, "material_profit": profit})

@estimate_bp.route('/api/items/<int:item_id>', methods=['DELETE'])
@login_required
@_require_csrf
def api_delete_item(item_id):
    n = execute_rowcount(
        "DELETE FROM estimate_items WHERE id=? AND estimate_id IN (SELECT id FROM estimates WHERE user_id=?)",
        (item_id, current_user.id),
    )
    if n == 0:
        return jsonify({"error": "Not found"}), 404
    return jsonify({"ok": True})

# ============================
# КАТАЛОГ
# ============================

def _add_to_catalog(section, name, unit, price, purchase):
    if not name: return
    try:
        if section == 'material':
            exists = fetch_one("SELECT id FROM catalog_materials WHERE name = ? AND user_id = ?", (name, current_user.id))
            if not exists:
                execute("""INSERT INTO catalog_materials 
                    (user_id, name, unit, category, purchase_price, retail_price, wholesale_price, min_wholesale_qty, description, use_count) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""", 
                    (current_user.id, name, unit, '', purchase, price, price, 10, ''))
            else:
                execute("UPDATE catalog_materials SET use_count = use_count + 1 WHERE name = ? AND user_id = ?", (name, current_user.id))
        else:
            exists = fetch_one("SELECT id FROM catalog_works WHERE name = ? AND user_id = ?", (name, current_user.id))
            if not exists:
                execute("""INSERT INTO catalog_works 
                    (user_id, name, unit, price, description, use_count) 
                    VALUES (?, ?, ?, ?, ?, 1)""", 
                    (current_user.id, name, unit, price, ''))
            else:
                execute("UPDATE catalog_works SET use_count = use_count + 1 WHERE name = ? AND user_id = ?", (name, current_user.id))
    except Exception as e:
        print(f"Catalog Error: {e}")

@estimate_bp.route('/api/catalog/materials', methods=['GET'])
@login_required
def api_get_materials():
    rows = fetch_all("SELECT * FROM catalog_materials WHERE user_id = ? ORDER BY use_count DESC", (current_user.id,))
    for row in rows:
        r = _safe_float(row.get('retail_price'))
        row['purchase_price'] = _normalize_unit_purchase_price(
            _safe_float(row.get('purchase_price')), r, qty=1.0
        )
    return jsonify(rows)

@estimate_bp.route('/api/catalog/materials', methods=['POST'])
@login_required
@_require_csrf
def api_add_catalog_material():
    data, err = _require_json()
    if err: return err
    try:
        retail = _safe_float(data.get('retail_price'))
        purchase = _normalize_unit_purchase_price(
            _safe_float(data.get('purchase_price')), retail, qty=1.0
        )
        wholesale = _safe_float(data.get('wholesale_price'), default=retail)
        category = data.get('category', '')
        min_qty = _safe_float(data.get('min_wholesale_qty'), default=10)
        desc = data.get('description', '')
        execute("""INSERT INTO catalog_materials
            (user_id, name, unit, category, article, brand, item_type, purchase_price, retail_price, wholesale_price, min_wholesale_qty, description, use_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
            (current_user.id, data.get('name', ''), data.get('unit', 'шт'), category,
             (data.get('article') or '').strip(), (data.get('brand') or '').strip(), (data.get('item_type') or '').strip(),
             purchase, retail, wholesale, min_qty, desc))
        return jsonify({"status": "ok"}), 201
    except Exception as e:
        logger.error(f"Add material error: {e}")
        return jsonify({"error": "Ошибка при создании материала"}), 400

@estimate_bp.route('/api/catalog/materials/<int:item_id>', methods=['PUT'])
@login_required
@_require_csrf
def api_update_catalog_material(item_id):
    data, err = _require_json()
    if err: return err
    try:
        retail = _safe_float(data.get('retail_price'))
        purchase = _normalize_unit_purchase_price(
            _safe_float(data.get('purchase_price')), retail, qty=1.0
        )
        execute("""UPDATE catalog_materials SET name=?, unit=?, category=?, article=?, brand=?, item_type=?,
                   purchase_price=?, retail_price=?, wholesale_price=?, min_wholesale_qty=?, description=?
                   WHERE id=? AND user_id=?""",
            (data.get('name'), data.get('unit'), data.get('category', ''),
             (data.get('article') or '').strip(), (data.get('brand') or '').strip(), (data.get('item_type') or '').strip(),
             purchase, retail,
             _safe_float(data.get('wholesale_price')), _safe_float(data.get('min_wholesale_qty'), default=10),
             data.get('description', ''), item_id, current_user.id))
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"Update material error: {e}")
        return jsonify({"error": "Ошибка при обновлении материала"}), 400

@estimate_bp.route('/api/catalog/materials/<int:item_id>', methods=['DELETE'])
@login_required
@_require_csrf
def api_delete_catalog_material(item_id):
    execute("DELETE FROM catalog_materials WHERE id = ? AND user_id = ?", (item_id, current_user.id))
    return jsonify({"ok": True})

@estimate_bp.route('/api/catalog/works', methods=['GET'])
@login_required
def api_get_works():
    return jsonify(fetch_all("SELECT * FROM catalog_works WHERE user_id = ? ORDER BY use_count DESC", (current_user.id,)))

@estimate_bp.route('/api/catalog/works', methods=['POST'])
@login_required
@_require_csrf
def api_add_catalog_work():
    data, err = _require_json()
    if err: return err
    try:
        price = _safe_float(data.get('price'))
        desc = data.get('description', '')
        execute("""INSERT INTO catalog_works
            (user_id, name, unit, price, description, use_count)
            VALUES (?, ?, ?, ?, ?, 1)""",
            (current_user.id, data.get('name', ''), data.get('unit', 'шт'), price, desc))
        return jsonify({"status": "ok"}), 201
    except Exception as e:
        logger.error(f"Add work error: {e}")
        return jsonify({"error": "Ошибка при создании работы"}), 400

@estimate_bp.route('/api/catalog/works/<int:item_id>', methods=['PUT'])
@login_required
@_require_csrf
def api_update_catalog_work(item_id):
    data, err = _require_json()
    if err: return err
    try:
        execute("""UPDATE catalog_works SET name=?, unit=?, price=?, description=?
                   WHERE id=? AND user_id=?""",
            (data.get('name'), data.get('unit'), _safe_float(data.get('price')), data.get('description', ''), item_id, current_user.id))
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"Update work error: {e}")
        return jsonify({"error": "Ошибка при обновлении работы"}), 400

@estimate_bp.route('/api/catalog/works/<int:item_id>', methods=['DELETE'])
@login_required
@_require_csrf
def api_delete_catalog_work(item_id):
    execute("DELETE FROM catalog_works WHERE id = ? AND user_id = ?", (item_id, current_user.id))
    return jsonify({"ok": True})

# ============================
# ЭКСПОРТ И СВЯЗИ
# ============================

@estimate_bp.route('/api/estimates/by-object/<int:object_id>', methods=['GET'])
@login_required
def api_get_estimates_by_object(object_id):
    estimates = fetch_all("SELECT * FROM estimates WHERE object_id = ? AND user_id = ?", (object_id, current_user.id))
    if not estimates:
        return jsonify({
            'estimates': [],
            'total_works': 0, 'total_materials': 0,
            'total_material_profit': 0
        })

    # Один запрос вместо N+1: загружаем все строки всех смет
    est_ids = [e['id'] for e in estimates]
    placeholders = ','.join('?' for _ in est_ids)
    all_items = fetch_all(
        f"SELECT * FROM estimate_items WHERE estimate_id IN ({placeholders})",
        tuple(est_ids)
    )
    # Группируем по estimate_id
    items_by_est = {}
    for item in all_items:
        items_by_est.setdefault(item['estimate_id'], []).append(item)

    total_works = 0; total_materials = 0; total_mat_profit = 0
    result_estimates = []
    for est in estimates:
        items = items_by_est.get(est['id'], [])
        est_works = sum(i['total'] for i in items if i['section'] == 'work')
        est_mats = sum(i['total'] for i in items if i['section'] == 'material')
        est_profit = sum(i.get('material_profit', 0) for i in items if i['section'] == 'material')
        total_works += est_works; total_materials += est_mats; total_mat_profit += est_profit
        sub = est_works + est_mats
        markup = sub * _safe_float(est.get('markup_percent')) / 100
        vat = (sub + markup) * _safe_float(est.get('vat_percent')) / 100
        disc = (sub + markup + vat) * _safe_float(est.get('discount_percent')) / 100
        result_estimates.append({
            'id': est['id'], 'number': est['number'], 'date': est['date'], 'status': est['status'],
            'total_works': est_works, 'total_materials': est_mats, 'total': sub + markup + vat - disc
        })
    return jsonify({
        'estimates': result_estimates,
        'total_works': total_works, 'total_materials': total_materials,
        'total_material_profit': total_mat_profit
    })

def _sanitize_excel(val):
    """Защита от формульных инъекций в Excel (CVE-2014-6352)"""
    if val is None:
        return ''
    s = str(val)
    if s.startswith(('=', '+', '-', '@')):
        s = "'" + s
    return s

@estimate_bp.route('/api/estimates/<int:est_id>/export', methods=['GET'])
@login_required
def api_export_excel(est_id):
    est = fetch_one("SELECT * FROM estimates WHERE id = ? AND user_id = ?", (est_id, current_user.id))
    if not est:
        return jsonify({"error": "Not found"}), 404

    items = fetch_all("SELECT * FROM estimate_items WHERE estimate_id = ? ORDER BY sort_order", (est_id,))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Смета {est['number']}"

    font_title = Font(name='Arial', size=14, bold=True, color="2E75B6")
    font_hdr = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    fill_hdr = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type='solid')
    fill_sec = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type='solid')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:F1')
    ws['A1'].value = f"Смета №{est['number']} от {est['date']}"
    ws['A1'].font = font_title

    ws['A3'].value = f"Объект: {_sanitize_excel(est['object_name'])}"
    ws['A4'].value = f"Клиент: {_sanitize_excel(est['client'])}"

    row = 6
    headers = ["№", "Наименование", "Ед.изм", "Кол-во", "Цена", "Сумма"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font_hdr
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal='center')
        c.border = thin

    mats = [x for x in items if x['section'] == 'material']
    works = [x for x in items if x['section'] == 'work']
    row = 7
    total_all = 0

    if mats:
        ws.merge_cells(f'A{row}:F{row}')
        c = ws.cell(row=row, column=1, value="МАТЕРИАЛЫ")
        c.font = font_hdr
        c.fill = fill_sec
        c.border = thin
        row += 1
        for i, it in enumerate(mats, 1):
            ws.cell(row=row, column=1, value=i).border = thin
            ws.cell(row=row, column=2, value=_sanitize_excel(it['name'])).border = thin
            ws.cell(row=row, column=3, value=_normalize_estimate_unit(it.get('unit'), 'material')).border = thin
            ws.cell(row=row, column=4, value=it['quantity']).border = thin
            ws.cell(row=row, column=5, value=it['price']).border = thin
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=it['total']).border = thin
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            total_all += it['total']
            row += 1

    if works:
        ws.merge_cells(f'A{row}:F{row}')
        c = ws.cell(row=row, column=1, value="РАБОТЫ")
        c.font = font_hdr
        c.fill = fill_sec
        c.border = thin
        row += 1
        for i, it in enumerate(works, 1):
            ws.cell(row=row, column=1, value=i).border = thin
            ws.cell(row=row, column=2, value=_sanitize_excel(it['name'])).border = thin
            ws.cell(row=row, column=3, value=_normalize_estimate_unit(it.get('unit'), 'work')).border = thin
            ws.cell(row=row, column=4, value=it['quantity']).border = thin
            ws.cell(row=row, column=5, value=it['price']).border = thin
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=it['total']).border = thin
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            total_all += it['total']
            row += 1

    ws.cell(row=row + 1, column=5, value="ИТОГО:").font = Font(bold=True, size=12)
    ws.cell(row=row + 1, column=6, value=total_all).font = Font(bold=True, size=12)
    ws.cell(row=row + 1, column=6).number_format = '#,##0.00'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Smeta_{est['number']}.xlsx"
    )
